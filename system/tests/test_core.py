"""اختبارات المحرك — البلوكات، العرض، الأمان، وواجهة المحرر."""

import json
import re

from django.contrib.auth.models import User
from django.test import TestCase
from django.utils import timezone
from datetime import timedelta

from system import blocks as B
from system.data import golden_classic
from system.models import (
        Asset, CustomFont, Customer, FavoriteBlock, Guest, Invitation, IntroVideo, MusicTrack, Order,
    OrderAddon, Plan, PlanAddon, RSVPResponse, SiteSetting, Template,

)
from system.forms import OrderForm, PlanAddonForm
from decimal import Decimal
from urllib.parse import quote
from system.renderer import block_style_css, layout_css, render_document

from system.sanitize import clean_html
from system import cssscope, customtext, guestimport, images, templateimport, video
from system.templatetags import invite as invite_tags
from django.core.files.uploadedfile import SimpleUploadedFile
from django.conf import settings
from pathlib import Path


# ==========================================================================
class SanitizeTests(TestCase):
    def test_removes_script_and_content(self):
        out = clean_html('مرحباً<script>alert(1)</script>بكم')
        self.assertNotIn("script", out)
        self.assertNotIn("alert", out)
        self.assertIn("مرحباً", out)

    def test_removes_event_handlers(self):
        out = clean_html('<p onclick="steal()">نص</p>')
        self.assertNotIn("onclick", out)
        self.assertIn("نص", out)

    def test_blocks_javascript_urls(self):
        out = clean_html('<a href="javascript:alert(1)">اضغط</a>')
        self.assertNotIn("javascript", out)

    def test_keeps_allowed_markup(self):
        out = clean_html("<p><strong>مهم</strong> وعادي</p>")
        self.assertIn("<strong>", out)

    def test_balances_unclosed_tags(self):
        out = clean_html("<p><strong>نص")
        self.assertTrue(out.endswith("</strong></p>"))

    def test_strips_dangerous_css(self):
        out = clean_html('<p style="color:red;background:url(evil)">نص</p>')
        self.assertIn("color:red", out)
        self.assertNotIn("url(", out)


# ==========================================================================
class DocumentTests(TestCase):
    def test_unknown_block_type_dropped(self):
        doc = B.normalize_document({"blocks": [{"type": "evil_block"}, {"type": "text"}]})
        self.assertEqual(len(doc["blocks"]), 1)
        self.assertEqual(doc["blocks"][0]["type"], "text")

    def test_singleton_enforced(self):
        doc = B.normalize_document({"blocks": [{"type": "hero"}, {"type": "hero"}]})
        self.assertEqual(len(doc["blocks"]), 1)

    def test_range_clamped_to_bounds(self):
        doc = B.normalize_document({"theme": {"radius": 99999}})
        self.assertLessEqual(doc["theme"]["radius"], 40)

    def test_invalid_color_falls_back(self):
        doc = B.normalize_document({"theme": {"accent": "url(javascript:1)"}})
        self.assertEqual(doc["theme"]["accent"], "#b8914f")

    def test_invalid_select_falls_back(self):
        doc = B.normalize_document({"theme": {"pattern": "../../etc/passwd"}})
        self.assertEqual(doc["theme"]["pattern"], "none")

    def test_javascript_url_rejected(self):
        doc = B.normalize_document({
            "blocks": [{"type": "location", "props": {"map_embed": "javascript:alert(1)"}}]
        })
        self.assertEqual(doc["blocks"][0]["props"]["map_embed"], "")

    def test_duplicate_ids_regenerated(self):
        doc = B.normalize_document({"blocks": [
            {"type": "text", "id": "same"}, {"type": "text", "id": "same"},
        ]})
        self.assertNotEqual(doc["blocks"][0]["id"], doc["blocks"][1]["id"])

    def test_block_limit_enforced(self):
        doc = B.normalize_document({"blocks": [{"type": "text"}] * 500})
        self.assertLessEqual(len(doc["blocks"]), 120)

    def test_every_registered_block_has_a_template(self):
        from django.template.loader import get_template
        for btype in B.BLOCK_REGISTRY:
            get_template(f"blocks/{btype}.html")   # يرمي استثناء لو مفقود


# ==========================================================================
class RendererTests(TestCase):
    def test_renders_every_block_type(self):
        doc = B.empty_document()
        doc["blocks"] = [B.make_block(t) for t in B.BLOCK_REGISTRY]
        out = render_document(doc)
        self.assertGreater(len(str(out["html"])), 500)
        self.assertIn("--accent", out["css_vars"])

    def test_hidden_block_not_sent_to_guest(self):
        doc = B.empty_document()
        doc["blocks"] = [B.make_block("text", props={"heading": "سرّي"}, visible=False)]
        out = render_document(doc, editable=False)
        self.assertNotIn("سرّي", str(out["html"]))

    def test_hidden_block_visible_in_editor(self):
        doc = B.empty_document()
        doc["blocks"] = [B.make_block("text", props={"heading": "سرّي"}, visible=False)]
        out = render_document(doc, editable=True)
        self.assertIn("سرّي", str(out["html"]))

    def test_feature_outside_plan_still_renders(self):
        doc = B.empty_document()
        doc["blocks"] = [B.make_block("gallery", props={"heading": "المعرض"})]
        out = render_document(doc, allowed_features={"rsvp"}, editable=False)
        self.assertIn("المعرض", str(out["html"]))

    def test_uploaded_font_is_emitted_as_font_face(self):
        font = CustomFont.objects.create(
            name="خط الدعوة", family="WeddingFont",
            file=SimpleUploadedFile("wedding.woff2", b"fake-font", content_type="font/woff2"),
        )
        out = render_document(B.empty_document())
        css = str(out["font_css"])
        self.assertIn("font-family:\"WeddingFont\"", css)
        self.assertIn(font.url, css)
        self.assertIn("format('woff2')", css)

    def test_xss_in_text_field_is_escaped(self):

        doc = B.empty_document()
        doc["blocks"] = [B.make_block("text", props={"heading": '<img src=x onerror=alert(1)>'})]
        html = str(render_document(doc)["html"])
        # يجب أن يظهر النص مهرَّباً لا كوسم فعلي
        self.assertNotIn("<img", html)
        self.assertIn("&lt;img", html)

    def test_golden_template_builds_and_renders(self):
        doc = golden_classic.build()
        self.assertGreater(len(doc["blocks"]), 8)
        self.assertGreater(len(str(render_document(doc)["html"])), 1000)


# ==========================================================================
class BaseAppTest(TestCase):
    def setUp(self):
        self.plan = Plan.objects.create(
            name="مميزة", slug="plus", price=1000,
            features=["rsvp", "countdown", "location", "qr", "gallery", "companions"],
        )
        self.basic = Plan.objects.create(
            name="أساسية", slug="basic", price=500, features=["countdown"],
        )
        self.template = Template.objects.create(
            name="ذهبي", slug="golden", document=golden_classic.build(),
        )
        self.customer = Customer.objects.create(name="عميل", phone="0100")
        self.inv = Invitation.objects.create(
            customer=self.customer, template=self.template, plan=self.plan,
            name_one="ليلى", name_two="أحمد", status="published",
            event_date=timezone.now() + timedelta(days=30),
            document=self.template.get_document(),
        )
        self.staff = User.objects.create_user("staff", password="Pass!12345x", is_staff=True)
        self.normal = User.objects.create_user("normal", password="Pass!12345x")
        # الكاش يعيش خارج المعاملة، فلا بد من تفريغه حتى لا يسرّب تحديد
        # المعدّل من اختبار إلى آخر.
        from django.core.cache import cache
        cache.clear()


# ==========================================================================
class AccessTests(BaseAppTest):
    def test_dashboard_requires_login(self):
        self.assertEqual(self.client.get("/dashboard/").status_code, 302)

    def test_dashboard_forbidden_for_non_staff(self):
        self.client.force_login(self.normal)
        self.assertEqual(self.client.get("/dashboard/").status_code, 403)

    def test_editor_forbidden_for_non_staff(self):
        self.client.force_login(self.normal)
        r = self.client.get(f"/dashboard/invitations/{self.inv.pk}/editor/")
        self.assertEqual(r.status_code, 403)

    def test_save_api_forbidden_for_non_staff(self):
        self.client.force_login(self.normal)
        r = self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/save/",
            data=json.dumps({"document": {}, "fields": {}}),
            content_type="application/json",
        )
        self.assertEqual(r.status_code, 403)

    def test_draft_invitation_is_404(self):
        self.inv.status = "draft"
        self.inv.save()
        self.assertEqual(self.client.get(f"/i/{self.inv.slug}/").status_code, 404)

    def test_expired_invitation_is_404(self):
        self.inv.expires_at = timezone.now() - timedelta(days=1)
        self.inv.save()
        self.assertEqual(self.client.get(f"/i/{self.inv.slug}/").status_code, 404)

    def test_password_protected_invitation(self):
        self.inv.password = "sirr"
        self.inv.save()
        r = self.client.get(f"/i/{self.inv.slug}/")
        self.assertContains(r, "محمية")
        self.client.post(f"/i/{self.inv.slug}/", {"password": "sirr"})
        self.assertContains(self.client.get(f"/i/{self.inv.slug}/"), "ليلى")


# ==========================================================================
class EditorApiTests(BaseAppTest):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.staff)

    def test_editor_page_loads(self):
        r = self.client.get(f"/dashboard/invitations/{self.inv.pk}/editor/")
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "editor-schema")
        self.assertContains(r, "editor-document")
        self.assertContains(r, "editor-fonts")

    def test_blank_template_page_loads(self):
        r = self.client.get("/dashboard/templates/new/")
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "إنشاء قالب جديد")
        self.assertContains(r, "إنشاء وفتح المحرر")

    def test_blank_template_creation_opens_editor(self):
        r = self.client.post("/dashboard/templates/new/", {
            "name": "قالب من الأقسام",
            "category": "wedding",
            "collection": "Premium",
            "description": "قالب مبني يدوياً",
            "slug": "ص",
        })
        self.assertEqual(r.status_code, 302)
        created = Template.objects.get(name="قالب من الأقسام")
        self.assertRedirects(r, f"/dashboard/templates/{created.pk}/editor/")
        self.assertEqual(created.source, "editor")
        self.assertEqual(created.created_by, self.staff)
        self.assertEqual(created.document["blocks"], [])
        self.assertIn("intro_play_mode", created.document["settings"])

    def test_blank_template_requires_a_name(self):
        before = Template.objects.count()
        r = self.client.post("/dashboard/templates/new/", {
            "name": "",
            "category": "wedding",
            "collection": "Premium",
        })
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "هذا الحقل مطلوب")
        self.assertEqual(Template.objects.count(), before)

    def test_preview_api_returns_html(self):

        r = self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/preview/",
            data=json.dumps({"document": self.inv.document}),
            content_type="application/json",
        )
        self.assertEqual(r.status_code, 200)
        data = r.json()
        self.assertTrue(data["ok"])
        self.assertIn("lb-", data["html"])

    def test_save_persists_document_and_fields(self):
        doc = B.empty_document()
        doc["blocks"] = [B.make_block("text", props={"heading": "عنوان محفوظ"})]
        r = self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/save/",
            data=json.dumps({"document": doc, "fields": {
                "name_one": "سارة", "name_two": "خالد", "status": "published",
                "event_type": "زفاف",
            }}),
            content_type="application/json",
        )
        self.assertEqual(r.status_code, 200)
        self.inv.refresh_from_db()
        self.assertEqual(self.inv.name_one, "سارة")
        self.assertEqual(self.inv.document["blocks"][0]["props"]["heading"], "عنوان محفوظ")

    def test_save_rejects_malicious_document(self):
        bad = {"blocks": [
            {"type": "text", "props": {"body": '<script>alert(1)</script>ok'}},
            {"type": "not_a_real_block"},
        ]}
        self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/save/",
            data=json.dumps({"document": bad, "fields": {"status": "draft", "event_type": "زفاف", "name_one": "ليلى"}}),
            content_type="application/json",
        )
        self.inv.refresh_from_db()
        self.assertEqual(len(self.inv.document["blocks"]), 1)
        # ملاحظة: نبحث عن الوسم نفسه لا عن كلمة "script"،
        # لأن اسم الحقل share_description يحتوي عليها.
        self.assertNotIn("<script", json.dumps(self.inv.document))
        self.assertNotIn("alert(1)", json.dumps(self.inv.document))
        self.assertEqual(self.inv.document["blocks"][0]["props"]["body"], "ok")

    def test_save_as_template_creates_template(self):
        r = self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/save-template/",
            data=json.dumps({"document": self.inv.document, "name": "قالب جديد"}),
            content_type="application/json",
        )
        self.assertEqual(r.status_code, 200)
        self.assertTrue(Template.objects.filter(name="قالب جديد").exists())

    def test_save_as_template_requires_name(self):
        r = self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/save-template/",
            data=json.dumps({"document": {}, "name": ""}),
            content_type="application/json",
        )
        self.assertEqual(r.status_code, 400)

    def test_upload_rejects_non_media_file(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        evil = SimpleUploadedFile("shell.php", b"<?php ?>", content_type="application/x-php")
        r = self.client.post(f"/dashboard/invitations/{self.inv.pk}/api/upload/", {"file": evil})
        self.assertEqual(r.status_code, 400)

    def test_delete_unused_image_from_library(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        asset = Asset.objects.create(
            file=SimpleUploadedFile("extra.png", b"not-used", content_type="image/png"),
            kind="image", original_name="extra.png", invitation=self.inv,
            uploaded_by=self.staff,
        )
        r = self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/assets/delete/",
            data=json.dumps({"asset": asset.pk}), content_type="application/json",
        )
        self.assertEqual(r.status_code, 200)
        self.assertFalse(Asset.objects.filter(pk=asset.pk).exists())

    def test_delete_used_image_is_rejected(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        asset = Asset.objects.create(
            file=SimpleUploadedFile("used.png", b"used", content_type="image/png"),
            kind="image", original_name="used.png", invitation=self.inv,
            uploaded_by=self.staff,
        )
        self.inv.document = {"blocks": [{"type": "text", "props": {"heading": asset.url}}]}
        self.inv.save(update_fields=["document"])
        r = self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/assets/delete/",
            data=json.dumps({"asset": asset.pk}), content_type="application/json",
        )
        self.assertEqual(r.status_code, 409)
        self.assertTrue(Asset.objects.filter(pk=asset.pk).exists())

    def test_bulk_delete_unused_images(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        assets = [Asset.objects.create(
            file=SimpleUploadedFile(f"extra-{i}.png", b"unused", content_type="image/png"),
            kind="image", original_name=f"extra-{i}.png", invitation=self.inv,
            uploaded_by=self.staff,
        ) for i in range(2)]
        r = self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/assets/bulk-delete/",
            data=json.dumps({"assets": [a.pk for a in assets]}),
            content_type="application/json",
        )
        self.assertEqual(r.status_code, 200)
        self.assertFalse(Asset.objects.filter(pk__in=[a.pk for a in assets]).exists())

    def test_bulk_delete_aborts_when_any_image_is_used(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        used = Asset.objects.create(
            file=SimpleUploadedFile("used-bulk.png", b"used", content_type="image/png"),
            kind="image", original_name="used-bulk.png", invitation=self.inv,
            uploaded_by=self.staff,
        )
        unused = Asset.objects.create(
            file=SimpleUploadedFile("unused-bulk.png", b"unused", content_type="image/png"),
            kind="image", original_name="unused-bulk.png", invitation=self.inv,
            uploaded_by=self.staff,
        )
        self.inv.document = {"blocks": [{"type": "text", "props": {"heading": used.url}}]}
        self.inv.save(update_fields=["document"])
        r = self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/assets/bulk-delete/",
            data=json.dumps({"assets": [used.pk, unused.pk]}),
            content_type="application/json",
        )
        self.assertEqual(r.status_code, 409)
        self.assertEqual(Asset.objects.filter(pk__in=[used.pk, unused.pk]).count(), 2)


# ========================================================================== 
class RsvpTests(BaseAppTest):

    def url(self):
        return f"/i/{self.inv.slug}/rsvp/"

    def test_rsvp_saved(self):
        self.client.post(self.url(), {"name": "منى", "status": "attending", "companions": "2"})
        self.assertEqual(RSVPResponse.objects.count(), 1)
        self.assertEqual(RSVPResponse.objects.first().companions, 2)

    def test_companions_clamped_to_block_max(self):
        self.client.post(self.url(), {"name": "منى", "status": "attending", "companions": "9999"})
        self.assertLessEqual(RSVPResponse.objects.first().companions, 5)

    def test_honeypot_blocks_bot(self):
        self.client.post(self.url(), {"name": "بوت", "website": "spam.com"})
        self.assertEqual(RSVPResponse.objects.count(), 0)

    def test_empty_name_rejected(self):
        self.client.post(self.url(), {"name": "", "status": "attending"})
        self.assertEqual(RSVPResponse.objects.count(), 0)

    def test_duplicate_within_window_ignored(self):
        for _ in range(3):
            self.client.post(self.url(), {"name": "منى", "status": "attending"})
        self.assertEqual(RSVPResponse.objects.count(), 1)

    def test_rate_limited(self):
        from django.core.cache import cache
        cache.clear()
        blocked = False
        for i in range(20):
            r = self.client.post(self.url(), {"name": f"ضيف {i}", "status": "attending"},
                                 HTTP_X_REQUESTED_WITH="XMLHttpRequest")
            if r.status_code == 429:
                blocked = True
                break
        self.assertTrue(blocked, "لم يُفعَّل تحديد المعدّل")

    def test_rsvp_works_when_plan_lacks_feature(self):
        self.inv.plan = self.basic
        self.inv.save()
        r = self.client.post(self.url(), {"name": "منى", "status": "attending"},
                             HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.json()["ok"])
        self.assertEqual(RSVPResponse.objects.count(), 1)


# ==========================================================================
class ViewCounterTests(BaseAppTest):
    def test_views_increment_without_race(self):
        for _ in range(5):
            self.client.get(f"/i/{self.inv.slug}/")
        self.inv.refresh_from_db()
        self.assertEqual(self.inv.public_views, 5)





# ==========================================================================
class AssetCacheBustTests(BaseAppTest):
    def test_invitation_html_has_content_versioned_assets_and_no_store(self):
        response = self.client.get(f"/i/{self.inv.slug}/")
        body = response.content.decode("utf-8")
        self.assertRegex(body, r"css/invite\.css\?v=[0-9a-f]{16}")
        self.assertRegex(body, r"js/invite\.js\?v=[0-9a-f]{16}")
        self.assertIn("no-store", response.get("Cache-Control", ""))


# ==========================================================================
class LayoutTests(BaseAppTest):

    """مواضع النصوص (السحب بالماوس) — تنظيف المدخلات وتوليد الـCSS."""

    def _layout(self, raw):
        doc = B.normalize_document({"blocks": [
            {"type": "hero", "id": "hero-1", "layout": raw}
        ]})
        return doc["blocks"][0]["layout"]

    def test_valid_offsets_kept(self):
        self.assertEqual(self._layout({"name_one": {"dx": 5.5, "dy": -3.25}}),
                         {"name_one": {"dx": 5.5, "dy": -3.25}})

    def test_out_of_range_is_clamped(self):
        got = self._layout({"name_one": {"dx": 9999, "dy": -9999}})
        self.assertEqual(got["name_one"],
                         {"dx": B.LAYOUT_MAX_X, "dy": -B.LAYOUT_MAX_Y})

    def test_zero_offset_is_dropped(self):
        self.assertEqual(self._layout({"name_one": {"dx": 0, "dy": 0}}), {})

    def test_unknown_and_malicious_slots_rejected(self):
        self.assertEqual(self._layout({
            "not_a_real_slot": {"dx": 3, "dy": 3},
            '"] {display:none} [x="': {"dx": 3, "dy": 3},
            "name_one": "مش قاموس",
            "subtitle": {"dx": "نص", "dy": None},
        }), {})

    def test_layout_css_only_emits_moved_slots(self):
        css = layout_css([
            {"id": "hero-1", "layout": {"name_one": {"dx": 4, "dy": -2},
                                        "subtitle": {"dx": 0, "dy": 0}}},
            {"id": "quote-1", "layout": {}},
        ])
        self.assertIn('#hero-1 [data-move="name_one"]{--dx:4cqw;--dy:-2cqw}', css)
        self.assertNotIn("subtitle", css)
        self.assertNotIn("quote-1", css)

    def test_layout_css_empty_when_nothing_moved(self):
        self.assertEqual(layout_css([{"id": "hero-1", "layout": {}}]), "")

    def test_section_height_is_sanitized_and_emitted(self):
        doc = B.normalize_document({"blocks": [{
            "type": "hero", "id": "hero-1", "style": {"section_height": 640}
        }]})
        style = doc["blocks"][0]["style"]
        self.assertEqual(style["section_height"], 640)
        css = block_style_css(style, B.default_theme())
        self.assertIn("--block-section-height:640px", css)



    def test_imported_section_height_and_background_are_rendered(self):
        doc = B.normalize_document({"blocks": [{
            "type": "custom_html", "id": "imp-1",
            "props": {"html": "<div class='countdown-grid'>محتوى</div>"},
            "style": {"section_height": 640, "bg_color": "#123456"},
        }]})
        rendered = render_document(doc, editable=False)
        html = str(rendered["html"])
        self.assertIn("lb--custom-bg", html)
        self.assertIn("--block-section-height:640px", html)

    def test_layout_renders_on_public_invitation(self):

        doc = self.inv.document
        doc["blocks"][0]["layout"] = {"name_one": {"dx": 7.5, "dy": 2.5}}
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])

        body = self.client.get(self.inv.get_absolute_url()).content.decode()
        self.assertIn("--dx:7.5cqw", body)
        # لازم يخرج غير مهرَّب، وإلا الـCSS تبقى غير صالحة تماماً
        self.assertNotIn("&quot;name_one&quot;", body)


# ==========================================================================
class GuestLinkTests(BaseAppTest):
    """الرابط الشخصي للضيف — الرمز هو بيانات الاعتماد."""

    def setUp(self):
        super().setUp()
        self.guest = Guest.objects.create(
            invitation=self.inv, name="أحمد سالم",
            phone="01000000001", plus_ones_allowed=2,
        )

    # ---- الوصول
    def test_guest_link_opens(self):
        res = self.client.get(self.guest.get_absolute_url())
        self.assertEqual(res.status_code, 200)
        self.assertIn(self.guest.name, res.content.decode())

    def test_form_is_prefilled_and_carries_token(self):
        body = self.client.get(self.guest.get_absolute_url()).content.decode()
        self.assertIn(f'name="guest_token" value="{self.guest.token}"', body)
        self.assertIn(f'value="{self.guest.name}"', body)

    def test_bad_token_is_404(self):
        self.assertEqual(
            self.client.get(f"/i/{self.inv.slug}/g/{'z' * 24}/").status_code, 404)

    def test_token_from_another_invitation_is_404(self):
        other = Invitation.objects.create(
            customer=self.customer, template=self.template, plan=self.plan,
            name_one="س", name_two="ص", status="published",
            event_date=timezone.now() + timedelta(days=10),
            document=self.template.get_document(),
        )
        stranger = Guest.objects.create(invitation=other, name="غريب")
        self.assertEqual(
            self.client.get(f"/i/{self.inv.slug}/g/{stranger.token}/").status_code, 404)

    def test_token_is_long_and_unique(self):
        other = Guest.objects.create(invitation=self.inv, name="ضيف تاني")
        self.assertGreaterEqual(len(self.guest.token), 20)
        self.assertNotEqual(self.guest.token, other.token)

    def test_draft_invitation_stays_hidden_even_with_token(self):
        self.inv.status = "draft"
        self.inv.save(update_fields=["status"])
        self.assertEqual(self.client.get(self.guest.get_absolute_url()).status_code, 404)

    # ---- الرد
    def test_rsvp_links_to_guest_and_caps_companions(self):
        self.client.post(f"/i/{self.inv.slug}/rsvp/", {
            "guest_token": self.guest.token, "name": self.guest.name,
            "status": "attending", "companions": "99",
        })
        r = RSVPResponse.objects.get(guest=self.guest)
        self.assertEqual(r.companions, self.guest.plus_ones_allowed)

    def test_guest_can_change_answer_without_duplicate(self):
        for status in ("attending", "declined"):
            self.client.post(f"/i/{self.inv.slug}/rsvp/", {
                "guest_token": self.guest.token, "name": self.guest.name,
                "status": status, "companions": "0",
            })
        self.assertEqual(RSVPResponse.objects.filter(guest=self.guest).count(), 1)
        self.assertEqual(RSVPResponse.objects.get(guest=self.guest).status, "declined")

    def test_anonymous_duplicate_is_still_blocked(self):
        for _ in range(2):
            self.client.post(f"/i/{self.inv.slug}/rsvp/",
                             {"name": "زائر مجهول", "status": "attending"})
        self.assertEqual(RSVPResponse.objects.filter(name="زائر مجهول").count(), 1)

    def test_forged_token_does_not_attach_to_an_existing_guest(self):
        """الرمز المزوّر مايخلّيش حد ينتحل شخصية ضيف موجود.

        الرد بيترتبط بسجل ضيف **جديد** باسم المُرسل (عشان ياخد تصريح
        دخوله)، مش بسجل الضيف اللي حاول ينتحله.
        """
        victim = self.guest
        self.client.post(f"/i/{self.inv.slug}/rsvp/", {
            "guest_token": "x" * 24, "name": "منتحل", "status": "attending",
        })
        rsvp = RSVPResponse.objects.get(name="منتحل")
        self.assertNotEqual(rsvp.guest_id, victim.pk)
        self.assertEqual(rsvp.guest.name, "منتحل")
        self.assertEqual(rsvp.guest.source, "rsvp")
        # وتصريح الضيف الأصلي ما اتلمسش
        victim.refresh_from_db()
        self.assertEqual(victim.entries_used, 0)


# ==========================================================================
class GuestImportTests(BaseAppTest):
    """استيراد CSV — الترميز هو اللي بيكسر الاستيراد في الواقع، مش المنطق."""

    CSV = ("الاسم,رقم الهاتف,المجموعة,عدد المرافقين\n"
           "أحمد سالم,01000123456,أهل العريس,2\n"
           "منى فؤاد,0100 123-4567,أهل العروسة,1\n")

    def test_reads_utf8_with_bom_from_excel(self):
        rows, rep = guestimport.parse_guests(self.CSV.encode("utf-8-sig"))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["name"], "أحمد سالم")

    def test_reads_windows_1256_from_arabic_excel(self):
        rows, rep = guestimport.parse_guests(self.CSV.encode("cp1256"))
        self.assertEqual(rep.encoding, "cp1256")
        self.assertEqual(rows[0]["name"], "أحمد سالم")

    def test_reads_semicolon_delimiter(self):
        rows, _ = guestimport.parse_guests(self.CSV.replace(",", ";").encode("utf-8"))
        self.assertEqual(len(rows), 2)

    def test_accepts_english_headers(self):
        rows, _ = guestimport.parse_guests(
            b"name,phone,group,plus\nAhmed,01001234567,Groom,2\n")
        self.assertEqual(rows[0]["name"], "Ahmed")
        self.assertEqual(rows[0]["plus_ones_allowed"], 2)

    def test_normalises_arabic_digits_and_spacing_in_phone(self):
        rows, _ = guestimport.parse_guests(
            "الاسم,رقم الهاتف\nسارة,٠١٠٠ ١٢٣-٤٥٦٧\n".encode("utf-8"))
        self.assertEqual(rows[0]["phone"], "01001234567")

    def test_missing_name_column_is_reported(self):
        rows, rep = guestimport.parse_guests("phone,city\n0100,القاهرة\n".encode())
        self.assertEqual(rows, [])
        self.assertTrue(any("عمود للاسم" in e for e in rep.errors))

    def test_rows_without_name_are_skipped(self):
        rows, rep = guestimport.parse_guests("الاسم,رقم الهاتف\n,0100\n".encode())
        self.assertEqual(rows, [])
        self.assertEqual(rep.skipped, 1)

    def test_oversized_file_is_refused(self):
        _, rep = guestimport.parse_guests(b"x" * (guestimport.MAX_BYTES + 1))
        self.assertTrue(rep.errors)

    # ---- الإدخال الفعلي
    def test_import_creates_guests_with_unique_tokens(self):
        rep = guestimport.import_guests(self.inv, self.CSV.encode("utf-8-sig"))
        self.assertEqual(rep.created, 2)
        tokens = set(self.inv.guests.values_list("token", flat=True))
        self.assertEqual(len(tokens), 2)
        self.assertTrue(all(len(t) >= 20 for t in tokens))

    def test_reimport_updates_instead_of_duplicating(self):
        guestimport.import_guests(self.inv, self.CSV.encode())
        changed = self.CSV.replace("أهل العريس,2", "أهل العريس,4")
        rep = guestimport.import_guests(self.inv, changed.encode())
        self.assertEqual(rep.created, 0)
        self.assertEqual(self.inv.guests.count(), 2)
        self.assertEqual(self.inv.guests.get(phone="01000123456").plus_ones_allowed, 4)

    def test_duplicate_phone_inside_one_file_is_skipped(self):
        raw = ("الاسم,رقم الهاتف\nأحمد,01000000009\nأحمد مرة تانية,01000000009\n")
        rep = guestimport.import_guests(self.inv, raw.encode())
        self.assertEqual(rep.created, 1)

    def test_upload_through_the_page(self):
        self.client.force_login(self.staff)
        upload = SimpleUploadedFile("guests.csv", self.CSV.encode("utf-8-sig"), "text/csv")
        res = self.client.post(f"/dashboard/invitations/{self.inv.pk}/guests/",
                               {"csv_file": upload})
        self.assertEqual(res.status_code, 302)
        self.assertEqual(self.inv.guests.count(), 2)


# ==========================================================================
class CheckinTests(BaseAppTest):
    """مسح الدخول على الباب."""

    def setUp(self):
        super().setUp()
        self.guest = Guest.objects.create(
            invitation=self.inv, name="ضيف مسح", plus_ones_allowed=3)
        self.url = f"/dashboard/invitations/{self.inv.pk}/checkin/scan/"

    def test_requires_staff(self):
        self.assertNotEqual(self.client.post(self.url, {"token": self.guest.token}).status_code, 200)

    def test_first_scan_marks_arrival(self):
        self.client.force_login(self.staff)
        data = self.client.post(self.url, {"token": self.guest.token}).json()
        self.assertTrue(data["ok"])
        self.assertFalse(data["already"])
        self.guest.refresh_from_db()
        self.assertTrue(self.guest.checked_in)
        self.assertIsNotNone(self.guest.checked_in_at)

    def test_second_scan_warns_instead_of_passing_silently(self):
        self.client.force_login(self.staff)
        self.client.post(self.url, {"token": self.guest.token})
        data = self.client.post(self.url, {"token": self.guest.token}).json()
        self.assertTrue(data["already"])

    def test_accepts_full_url_from_the_camera(self):
        self.client.force_login(self.staff)
        full = f"https://x.test/i/{self.inv.slug}/g/{self.guest.token}/"
        self.assertTrue(self.client.post(self.url, {"token": full}).json()["ok"])

    def test_unknown_and_short_tokens(self):
        self.client.force_login(self.staff)
        self.assertEqual(self.client.post(self.url, {"token": "z" * 24}).status_code, 404)
        self.assertEqual(self.client.post(self.url, {"token": "abc"}).status_code, 400)

    def test_token_from_another_invitation_is_rejected(self):
        other = Invitation.objects.create(
            customer=self.customer, template=self.template, plan=self.plan,
            name_one="س", name_two="ص", status="published",
            event_date=timezone.now() + timedelta(days=5),
            document=self.template.get_document(),
        )
        stranger = Guest.objects.create(invitation=other, name="غريب")
        self.client.force_login(self.staff)
        self.assertEqual(self.client.post(self.url, {"token": stranger.token}).status_code, 404)

    def test_guest_qr_is_served(self):
        res = self.client.get(f"/i/{self.inv.slug}/g/{self.guest.token}/qr.svg")
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["Content-Type"], "image/svg+xml")


# ==========================================================================
class FluidSizeTests(BaseAppTest):
    """المقاسات المرنة — تُضبط مرة على الديسكتوب وتتقلّص لوحدها."""

    def test_produces_clamp_with_container_units(self):
        out = invite_tags._fluid(72, ref=760)
        self.assertTrue(out.startswith("clamp("))
        self.assertIn("cqw", out)
        self.assertIn("72.0px", out)          # الحد الأقصى = اللي المستخدم كتبه

    def test_user_mobile_value_becomes_the_floor(self):
        self.assertTrue(invite_tags._fluid(72, floor=46).startswith("clamp(46.0px"))

    def test_floor_never_exceeds_the_desktop_size(self):
        out = invite_tags._fluid(30, floor=90)
        self.assertTrue(out.startswith("clamp(30.0px"))

    def test_zero_and_garbage_are_safe(self):
        for bad in ("", None, "abc", 0, -5):
            self.assertEqual(invite_tags._fluid(bad), "0px")

    def test_scales_relative_to_stage_width(self):
        # ٧٦ من ٧٦٠ = ١٠٪ من عرض المسرح
        self.assertIn("10.0cqw", invite_tags._fluid(76, ref=760))

    def test_rendered_invitation_uses_clamp_not_fixed_px(self):
        body = self.client.get(self.inv.get_absolute_url()).content.decode()
        self.assertIn("--name-size:clamp(", body)
        self.assertIn("--block-pt:clamp(", body)
        self.assertNotIn("--name-size-m:", body)   # الحقل القديم اتحوّل لحد أدنى


# ==========================================================================
class ImagePipelineTests(BaseAppTest):
    """ضغط الصور والقص — أهم بند لأن ناتجه بيوصل للضيوف."""

    def _photo(self, size=(4032, 3024)):
        from PIL import Image
        import io as _io
        img = Image.new("RGB", size, (180, 150, 120))
        buf = _io.BytesIO()
        img.save(buf, "JPEG", quality=92)
        return buf.getvalue()

    def _upload(self, raw, name="photo.jpg", ctype="image/jpeg"):
        self.client.force_login(self.staff)
        return self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/upload/",
            {"file": SimpleUploadedFile(name, raw, ctype)},
        )

    def test_large_photo_is_downscaled_and_converted(self):
        raw = self._photo()
        self.assertTrue(self._upload(raw).json()["ok"])
        asset = Asset.objects.latest("id")
        self.assertLessEqual(max(asset.width, asset.height), images.MAX_EDGE)
        self.assertTrue(asset.file.name.endswith(".webp"))
        self.assertLess(asset.size_bytes, len(raw) / 4)   # توفير معتبر مش تجميلي

    def test_thumbnail_and_source_are_kept(self):
        self._upload(self._photo((1200, 900)))
        asset = Asset.objects.latest("id")
        self.assertTrue(asset.thumb)
        self.assertTrue(asset.source)          # الأصل لازم يفضل عشان إعادة القص

    def test_small_image_is_not_upscaled(self):
        self._upload(self._photo((400, 300)))
        asset = Asset.objects.latest("id")
        self.assertEqual((asset.width, asset.height), (400, 300))

    def test_corrupt_file_is_refused(self):
        res = self._upload(b"not an image at all", "x.jpg")
        self.assertEqual(res.status_code, 400)

    # ---- القص
    def _crop(self, asset, box):
        return self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/crop/",
            data=json.dumps({"asset": asset.pk, "box": box}),
            content_type="application/json",
        )

    def test_crop_respects_the_requested_ratio(self):
        self._upload(self._photo())
        asset = Asset.objects.latest("id")
        data = self._crop(asset, {"x": 0, "y": 0, "w": 0.25, "h": 0.5}).json()
        self.assertTrue(data["ok"])
        got = data["asset"]["width"] / data["asset"]["height"]
        want = (0.25 * 4032) / (0.5 * 3024)
        self.assertAlmostEqual(got, want, delta=0.05)

    def test_crop_keeps_the_original_so_it_can_be_recropped(self):
        self._upload(self._photo())
        first = Asset.objects.latest("id")
        cropped = Asset.objects.get(pk=self._crop(first, {"x": 0, "y": 0, "w": .5, "h": .5}).json()["asset"]["id"])
        self.assertTrue(cropped.source)
        again = self._crop(cropped, {"x": 0, "y": 0, "w": .5, "h": .5})
        self.assertTrue(again.json()["ok"])

    def test_out_of_bounds_box_is_clamped_not_crashed(self):
        self._upload(self._photo((800, 600)))
        asset = Asset.objects.latest("id")
        res = self._crop(asset, {"x": 5, "y": 5, "w": 9, "h": 9})
        self.assertEqual(res.status_code, 200)

    def test_crop_requires_staff(self):
        self._upload(self._photo((400, 300)))
        asset = Asset.objects.latest("id")
        self.client.logout()
        self.assertNotEqual(self._crop(asset, {"x": 0, "y": 0, "w": 1, "h": 1}).status_code, 200)

    def test_crop_rejects_asset_from_another_invitation(self):
        other = Invitation.objects.create(
            customer=self.customer, template=self.template, plan=self.plan,
            name_one="س", name_two="ص", status="published",
            event_date=timezone.now() + timedelta(days=5),
            document=self.template.get_document(),
        )
        self.client.force_login(self.staff)
        stranger = Asset.objects.create(kind="image", invitation=other,
                                        original_name="x", width=10, height=10)
        self.assertEqual(self._crop(stranger, {"x": 0, "y": 0, "w": 1, "h": 1}).status_code, 403)


# ==========================================================================
class IntroTests(BaseAppTest):
    """الشاشة الافتتاحية — لازم تظهر في المحرر عشان تتعاين وتتعدّل."""

    def _enable(self, **extra):
        doc = self.inv.document
        doc["settings"]["intro_enabled"] = True
        doc["settings"].update(extra)
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])

    def test_intro_shows_for_guests(self):
        self._enable()
        self.assertIn("lb-intro", self.client.get(self.inv.get_absolute_url()).content.decode())

    def test_intro_shows_inside_the_editor_preview(self):
        self._enable()
        self.client.force_login(self.staff)
        body = self.client.get(
            f"/dashboard/invitations/{self.inv.pk}/preview-frame/").content.decode()
        self.assertIn("lb-intro", body)
        # العلامة دي بتخلي المحرر يقدر يرجّعها بعد ما تضغط «التالي»
        self.assertIn("data-intro-editable", body)

    def test_intro_video_is_muted_and_inline(self):
        self._enable(intro_video="/media/x.mp4")
        body = self.client.get(self.inv.get_absolute_url()).content.decode()
        self.assertIn("data-intro-video", body)
        # المتصفحات بتمنع التشغيل التلقائي بصوت، وiOS بيحتاج playsinline
        self.assertIn("muted", body)
        self.assertIn("playsinline", body)

    def test_intro_video_can_start_with_sound(self):
        self._enable(intro_video="/media/x.mp4", intro_video_audio="sound",
                     intro_play_mode="button")
        body = self.client.get(self.inv.get_absolute_url()).content.decode()
        tag = body[body.index("<video"):body.index("</video>")]
        self.assertIn('data-intro-audio="sound"', tag)
        self.assertNotIn(" muted", tag)
        self.assertNotIn("data-intro-sound", body)

    def test_intro_video_can_start_silent(self):
        self._enable(intro_video="/media/x.mp4", intro_video_audio="silent",
                     intro_play_mode="button")
        body = self.client.get(self.inv.get_absolute_url()).content.decode()
        tag = body[body.index("<video"):body.index("</video>")]
        self.assertIn('data-intro-audio="silent"', tag)
        self.assertIn("muted", tag)
        self.assertNotIn("data-intro-sound", body)

    def test_no_intro_markup_when_disabled(self):

        doc = self.inv.document
        doc["settings"]["intro_enabled"] = False
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])
        self.assertNotIn('class="lb-intro',
                         self.client.get(self.inv.get_absolute_url()).content.decode())


# ==========================================================================
class VideoUploadTests(BaseAppTest):
    """رفع فيديو الافتتاحية — كان بيترفض قبل كده لأن المُنتقي ما كانش بيقبله."""

    def _clip(self, seconds=20, height=1080):
        """يولّد MP4 حقيقي بـffmpeg. لو مش متثبّت بنتخطّى الاختبار."""
        import shutil, subprocess, tempfile, os
        if not shutil.which("ffmpeg"):
            self.skipTest("ffmpeg غير متاح")
        path = tempfile.mktemp(suffix=".mp4")
        subprocess.run(
            ["ffmpeg", "-y", "-loglevel", "error", "-f", "lavfi",
             # libx264 مع yuv420p بيرفض الأبعاد الفردية
             "-i", f"testsrc=size={height * 16 // 9 // 2 * 2}x{height}:rate=25:duration={seconds}",
             "-c:v", "libx264", "-preset", "ultrafast", "-pix_fmt", "yuv420p", path],
            check=True, timeout=120,
        )
        raw = open(path, "rb").read()
        os.unlink(path)
        return raw

    def _upload(self, raw, name="intro.mp4", ctype="video/mp4"):
        self.client.force_login(self.staff)
        return self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/upload/",
            {"file": SimpleUploadedFile(name, raw, ctype)},
        )

    def test_video_upload_is_accepted(self):
        data = self._upload(self._clip(seconds=3, height=360)).json()
        self.assertTrue(data["ok"])
        self.assertEqual(data["asset"]["kind"], "video")

    def test_tall_video_is_shrunk_to_720p(self):
        raw = self._clip(seconds=6, height=1080)
        data = self._upload(raw).json()
        asset = Asset.objects.latest("id")
        self.assertLess(asset.size_bytes, len(raw))          # اتضغط فعلاً
        self.assertGreater(data["asset"]["seconds"], 0)

    def test_duration_is_not_silently_trimmed(self):
        """القص الصامت بيبوّظ مقطع فرح طويل من غير ما المستخدم يعرف."""
        data = self._upload(self._clip(seconds=14, height=480)).json()
        self.assertGreater(data["asset"]["seconds"], 13)

    def test_audio_is_kept(self):
        """قسم الفيديو ممكن يكون مقطع ليه صوت — مانشيلوش."""
        import shutil, subprocess, tempfile, os
        if not shutil.which("ffmpeg"):
            self.skipTest("ffmpeg غير متاح")
        path = tempfile.mktemp(suffix=".mp4")
        subprocess.run(
            ["ffmpeg", "-y", "-loglevel", "error",
             "-f", "lavfi", "-i", "testsrc=size=640x480:rate=25:duration=4",
             "-f", "lavfi", "-i", "sine=frequency=440:duration=4",
             "-c:v", "libx264", "-preset", "ultrafast", "-pix_fmt", "yuv420p",
             "-c:a", "aac", "-shortest", path],
            check=True, timeout=120)
        raw = open(path, "rb").read()
        os.unlink(path)
        self.assertTrue(self._upload(raw).json()["ok"])
        asset = Asset.objects.latest("id")
        probe = subprocess.run(
            ["ffprobe", "-v", "error", "-select_streams", "a",
             "-show_entries", "stream=codec_type", "-of", "default=nw=1:nk=1",
             asset.file.path],
            capture_output=True, text=True, timeout=20)
        self.assertIn("audio", probe.stdout)

    def test_compress_falls_back_when_ffmpeg_is_missing(self):
        """الاستضافة ممكن تكون من غير ffmpeg — لازم الرفع يفضل شغّال."""
        original = video.available
        video.available = lambda: False
        try:
            up = SimpleUploadedFile("x.mp4", b"fake bytes", "video/mp4")
            out, secs = video.compress(up)
            self.assertIs(out, up)
            self.assertEqual(secs, 0)
        finally:
            video.available = original

    def test_disguised_file_is_refused(self):
        self.assertEqual(self._upload(b"x", "a.exe", "application/x-msdownload").status_code, 400)


# ==========================================================================
class MusicLibraryTests(BaseAppTest):
    """مكتبة الموسيقى — ترفع المقطوعة مرة وتختارها في أي دعوة."""

    def setUp(self):
        super().setUp()
        self.track = MusicTrack.objects.create(
            name="زفة كلاسيك", external_url="https://cdn.example.com/a.mp3")

    def test_page_requires_staff(self):
        self.assertNotEqual(self.client.get("/dashboard/music/").status_code, 200)

    def test_page_lists_tracks(self):
        self.client.force_login(self.staff)
        self.assertIn("زفة كلاسيك", self.client.get("/dashboard/music/").content.decode())

    def test_track_needs_a_file_or_a_url(self):
        from system.forms import MusicTrackForm
        self.assertFalse(MusicTrackForm({"name": "بدون صوت", "order": 0}).is_valid())
        self.assertTrue(MusicTrackForm(
            {"name": "برابط", "external_url": "https://x.test/a.mp3", "order": 0}).is_valid())

    def test_library_reaches_the_editor(self):
        self.client.force_login(self.staff)
        body = self.client.get(
            f"/dashboard/invitations/{self.inv.pk}/editor/").content.decode()
        self.assertIn("editor-music", body)
        # json_script بيهرّب العربي لـ\\uXXXX، فبندوّر على الشكل ده مش على الحروف
        self.assertIn(r"\u0632\u0641\u0629", body)

    def test_hidden_tracks_do_not_reach_the_editor(self):
        MusicTrack.objects.update(is_active=False)
        self.client.force_login(self.staff)
        body = self.client.get(
            f"/dashboard/invitations/{self.inv.pk}/editor/").content.decode()
        self.assertIn('id="editor-music" type="application/json">[]', body)

    def test_deleting_a_track(self):
        self.client.force_login(self.staff)
        self.client.post("/dashboard/music/",
                         {"action": "delete", "track": self.track.pk})
        self.assertFalse(MusicTrack.objects.filter(pk=self.track.pk).exists())

    def test_music_url_field_is_an_audio_picker(self):
        """قبل كده كان حقل نص — المستخدم كان لازم يعرف رابط جاهز."""
        field = next(f for f in B.editor_schema()["settings_fields"]
                     if f["key"] == "music_url")
        self.assertEqual(field["type"], "media")
        self.assertEqual(field["media_kind"], "audio")


# ==========================================================================
class CssScopeTests(TestCase):
    """حصر CSS — ده حد أمان مش تجميل، فلازم يتغطّى بالتفصيل."""

    def scope(self, css, url_map=None):
        return cssscope.scope_css(css, "#b1", url_map)

    def test_plain_rule_is_scoped(self):
        self.assertEqual(self.scope(".t{color:red}"), "#b1 .t{color:red}")

    def test_page_wide_selectors_become_the_block(self):
        for sel in ("html", "body", ":root", "*"):
            self.assertEqual(self.scope(sel + "{margin:0}"), "#b1{margin:0}")

    def test_body_descendant_is_rebased(self):
        self.assertEqual(self.scope("body .t{color:red}"), "#b1 .t{color:red}")

    def test_duplicate_selectors_are_merged(self):
        self.assertEqual(self.scope("html, body{margin:0}"), "#b1{margin:0}")

    def test_media_query_contents_are_scoped(self):
        out = self.scope("@media (max-width:600px){.t{color:red}}")
        self.assertIn("@media (max-width:600px){#b1 .t{color:red}}", out)

    def test_keyframes_pass_through_unscoped(self):
        out = self.scope("@keyframes fade{from{opacity:0}to{opacity:1}}")
        self.assertIn("@keyframes fade{", out)
        self.assertNotIn("#b1 from", out)

    def test_import_is_dropped_without_eating_the_next_rule(self):
        """الخطأ ده بلع القاعدة اللي بعده لما اتكتب أول مرة."""
        out = self.scope('@import url("//evil.test/x.css");\nbody{margin:0}')
        self.assertNotIn("evil.test", out)
        self.assertIn("#b1{margin:0}", out)

    def test_dangerous_declarations_are_dropped(self):
        css = ".t{width:expression(alert(1));behavior:url(x.htc);-moz-binding:url(y);color:red}"
        out = self.scope(css)
        self.assertNotIn("expression", out)
        self.assertNotIn("behavior", out)
        self.assertNotIn("binding", out)
        self.assertIn("color:red", out)

    def test_fixed_becomes_absolute(self):
        """fixed بيهرب من القسم ويفضل معلّق فوق باقي الدعوة."""
        self.assertIn("position:absolute", self.scope(".t{position:fixed}"))

    def test_relative_urls_map_to_stored_assets(self):
        out = self.scope(".t{background:url(bg.jpg)}", {"bg.jpg": "/media/a/bg.webp"})
        self.assertIn('url("/media/a/bg.webp")', out)

    def test_unmapped_relative_url_is_neutralised(self):
        self.assertIn("none", self.scope(".t{background:url(ghost.jpg)}"))

    def test_absolute_paths_survive_a_second_pass(self):
        """الحصر بيتعمل وقت العرض على CSS اتحلّت روابطه وقت الاستيراد."""
        once = self.scope(".t{background:url(bg.jpg)}", {"bg.jpg": "/media/a/bg.webp"})
        twice = cssscope.scope_css(once.replace("#b1 ", ""), "#b1")
        self.assertIn("/media/a/bg.webp", twice)

    def test_unbalanced_braces_do_not_crash(self):
        self.assertIsInstance(self.scope(".a{color:red}}}.b{color:blue"), str)


# ==========================================================================
class TemplateImportTests(TestCase):
    """استيراد قالب من ملف — الملف جاي من بره فالأمان أهم بند."""

    def _zip(self, files: dict):
        import io as _io, zipfile
        buf = _io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
            for name, data in files.items():
                z.writestr(name, data)
        return SimpleUploadedFile("t.zip", buf.getvalue(), "application/zip")

    PAGE = """<!doctype html><html><head><title>عمر و ياسمين</title>
      <link rel="stylesheet" href="s.css"><script src="x.js"></script>
      <style>.hero h1{font-size:64px}</style></head>
      <body onload="track()">
        <header class="hero"><h1>عمر و ياسمين</h1>
          <p>يتشرفان بدعوتكم لحضور حفل زفافهما بقاعة الماسة</p></header>
        <section class="story"><h2>قصتنا</h2>
          <p>اتقابلنا في الجامعة سنة ألفين وتسعتاشر وكمّلنا مع بعض من ساعتها.</p>
          <script>steal()</script></section>
      </body></html>"""
    CSS = "body{background:#111}.hero{min-height:100vh}"

    def test_plain_html_file_imports(self):
        up = SimpleUploadedFile("t.html", self.PAGE.encode(), "text/html")
        tpl = templateimport.import_template(up)
        self.assertEqual(len(tpl.document["blocks"]), 2)
        self.assertEqual(tpl.source, "import")

    def test_title_is_used_as_the_name(self):
        up = SimpleUploadedFile("t.html", self.PAGE.encode(), "text/html")
        self.assertEqual(templateimport.import_template(up).name, "عمر و ياسمين")

    def test_explicit_name_wins(self):
        up = SimpleUploadedFile("t.html", self.PAGE.encode(), "text/html")
        self.assertEqual(templateimport.import_template(up, name="اسمي").name, "اسمي")

    def test_scripts_and_handlers_are_stripped_from_editable_html(self):
        up = SimpleUploadedFile("t.html", self.PAGE.encode(), "text/html")
        tpl = templateimport.import_template(up)
        blob = json.dumps(tpl.document, ensure_ascii=False)
        self.assertNotIn("steal", blob)
        self.assertNotIn("onload", blob)
        self.assertNotIn("<script", blob)

    def test_local_scripts_are_saved_as_runtime_metadata(self):
        page = "<html><head><script src='js/app.js'></script></head><body><h1>دعوة</h1><p>محتوى كافٍ للمعاينة والتجربة.</p></body></html>"
        tpl = templateimport.import_template(
            self._zip({"index.html": page, "js/app.js": "window.__templateLoaded = true;"}),
            name="قالب JavaScript",
        )
        self.assertEqual(len(tpl.runtime_scripts), 1)
        self.assertTrue(tpl.runtime_scripts[0]["src"].startswith("/media/"))
        blob = json.dumps(tpl.document, ensure_ascii=False).lower()
        self.assertNotIn("window.__templateloaded", blob)
        self.assertNotIn("<script", blob)
        body = self.client.get(f"/templates/{tpl.slug}/preview/").content.decode()
        self.assertIn(tpl.runtime_scripts[0]["src"], body)



    def test_full_runtime_keeps_intro_and_embedded_iframe(self):
        page = (
            '<html><head><script src="js/app.js"></script></head><body>'
            '<div id="weiOverlay"><video id="weiVideo"></video></div>'
            '<section class="countdown"><div id="countdownContainer">00:00:00</div>'
            '<iframe src="embed.html" title="Map"></iframe></section>'
            '<p>محتوى كافٍ لاختبار القالب الكامل وتشغيل السكربتات.</p>'
            '</body></html>'
        )
        tpl = templateimport.import_template(
            self._zip({
                "index.html": page,
                "js/app.js": "window.__fullRuntime = true;",
                "embed.html": '<html><body><div id="mapDiv"></div><script src="js/map.js"></script></body></html>',
                "js/map.js": "window.__mapRuntime = true;",
            }),
            name="قالب Runtime كامل",
        )
        blob = json.dumps(tpl.document, ensure_ascii=False)
        self.assertIn("weiOverlay", blob)
        self.assertIn("countdownContainer", blob)
        self.assertIn("/media/template_pages/", blob)
        self.assertEqual(len(tpl.runtime_scripts), 1)
        page_html = self.client.get(f"/templates/{tpl.slug}/preview/").content.decode()
        self.assertIn("weiOverlay", page_html)
        self.assertIn("countdownContainer", page_html)
        self.assertIn(tpl.runtime_scripts[0]["src"], page_html)

    def test_expired_countdown_stays_visible_as_zero(self):
        page = (
            '<html><body><div id="countdownContainer">'
            '<div id="days">168</div><div id="hours">00</div>'
            '</div><script>'
            "var dist=eventDate-now;if(dist<0) {document.getElementById('countdownContainer').innerHTML='See you there!';return;}"
            '</script><p>محتوى كافٍ لاختبار العدّاد المنتهي.</p></body></html>'
        )
        tpl = templateimport.import_template(
            SimpleUploadedFile("countdown.html", page.encode(), "text/html"),
            name="عداد منتهي",
        )
        code = "\n".join(s.get("code") or "" for s in tpl.runtime_scripts)
        self.assertIn("if (dist < 0) { dist = 0; }", code)
        self.assertNotIn("See you there!", code)
        self.assertIn("countdownContainer", tpl.document["blocks"][0]["props"]["html"])

    def test_linked_stylesheet_is_picked_up(self):

        tpl = templateimport.import_template(
            self._zip({"index.html": self.PAGE, "s.css": self.CSS}))
        self.assertIn("min-height:100vh", tpl.document["blocks"][0]["props"]["css"])

    def test_stored_css_is_scoped_when_rendered(self):
        tpl = templateimport.import_template(
            self._zip({"index.html": self.PAGE, "s.css": self.CSS}))
        html = render_document(tpl.document, invitation=None, request=None,
                               allowed_features=None, editable=False)["html"]
        bid = tpl.document["blocks"][0]["id"]
        self.assertIn(f"#{bid}", html)
        # body{} اللي كان بيصبغ الصفحة كلها بقى على القسم نفسه
        self.assertNotIn("body{background", html.replace(" ", ""))

    def test_images_are_stored_and_relinked(self):
        from PIL import Image
        import io as _io
        buf = _io.BytesIO()
        Image.new("RGB", (900, 600), (1, 2, 3)).save(buf, "PNG")
        target = "<h1>عمر و ياسمين</h1>"
        self.assertIn(target, self.PAGE)      # لو الفيكستشر اتغيّر، نعرف فوراً
        page = self.PAGE.replace(target, target + '<img src="p.png" alt="">')
        tpl = templateimport.import_template(
            self._zip({"index.html": page, "p.png": buf.getvalue()}))
        blob = json.dumps(tpl.document)
        self.assertIn("/media/", blob)
        self.assertNotIn('src="p.png"', blob)
        self.assertTrue(Asset.objects.filter(invitation__isnull=True).exists())

    # ---- الأمان
    def test_path_traversal_members_are_ignored(self):
        up = self._zip({"../../../etc/passwd.html": "<body><p>x</p></body>",
                        "index.html": self.PAGE})
        tpl = templateimport.import_template(up)
        self.assertIn("عمر", json.dumps(tpl.document, ensure_ascii=False))

    def test_absolute_path_member_is_ignored(self):
        main, files = templateimport.parse_upload(
            self._zip({"/etc/x.html": "<body><p>x</p></body>", "index.html": self.PAGE}))
        self.assertNotIn("/etc/x.html", files)

    def test_unsupported_members_are_dropped_but_js_is_kept_for_runtime(self):
        _, files = templateimport.parse_upload(
            self._zip({"index.html": self.PAGE, "x.js": "steal()",
                       "x.php": "<?php ?>", "x.exe": "MZ"}))
        self.assertEqual(set(files), {"index.html", "x.js"})

    def test_zip_bomb_is_refused(self):
        bomb = self._zip({"index.html": self.PAGE, "big.css": "a" * (30 * 1024 * 1024)})
        with self.assertRaises(templateimport.ImportError_):
            templateimport.parse_upload(bomb)

    def test_archive_without_html_is_refused(self):
        with self.assertRaises(templateimport.ImportError_):
            templateimport.parse_upload(self._zip({"a.css": "x{}"}))

    def test_broken_zip_is_refused(self):
        up = SimpleUploadedFile("t.zip", b"PK\x03\x04 garbage", "application/zip")
        with self.assertRaises(templateimport.ImportError_):
            templateimport.parse_upload(up)

    def test_wrong_extension_is_refused(self):
        with self.assertRaises(templateimport.ImportError_):
            templateimport.parse_upload(
                SimpleUploadedFile("a.txt", b"hello", "text/plain"))

    def test_index_html_is_preferred_over_other_pages(self):
        main, _ = templateimport.parse_upload(self._zip({
            "pages/about.html": "<body><p>about</p></body>",
            "index.html": self.PAGE,
        }))
        self.assertEqual(main, "index.html")

    def test_slug_collision_gets_a_suffix(self):
        for _ in range(2):
            templateimport.import_template(
                SimpleUploadedFile("t.html", self.PAGE.encode(), "text/html"))
        slugs = list(Template.objects.filter(source="import").values_list("slug", flat=True))
        self.assertEqual(len(set(slugs)), len(slugs))


# ==========================================================================
class TemplateImportViewTests(BaseAppTest):
    def test_upload_requires_staff(self):
        self.assertNotEqual(self.client.get("/dashboard/templates/").status_code, 200)

    def test_upload_creates_a_template(self):
        self.client.force_login(self.staff)
        page = ("<html><head><title>Imported</title></head><body>"
                "<h1>عمر و ياسمين</h1>"
                "<p>يتشرفان بدعوتكم لحضور حفل زفافهما بقاعة الماسة يوم الجمعة</p>"
                "</body></html>").encode()
        res = self.client.post("/dashboard/templates/", {
            "template_file": SimpleUploadedFile("a.html", page, "text/html"),
        }, follow=True)
        self.assertEqual(res.status_code, 200)
        self.assertTrue(Template.objects.filter(source="import").exists())

    def test_bad_upload_shows_an_arabic_error_not_a_500(self):
        self.client.force_login(self.staff)
        res = self.client.post("/dashboard/templates/", {
            "template_file": SimpleUploadedFile("a.txt", b"nope", "text/plain"),
        }, follow=True)
        self.assertEqual(res.status_code, 200)
        self.assertFalse(Template.objects.filter(source="import").exists())


# ==========================================================================
class IntroLivePreviewTests(BaseAppTest):
    """الافتتاحية لازم تتحدّث لحظياً في المحرر.

    كانت مكتوبة جوّه render.html، والمحرر بيبدّل ‎.lb-stage‎ بس — والافتتاحية
    أخت للـstage مش جواه. فأي تعديل فيها ما كانش بيبان غير بعد ما تقفل
    المحرر وتفتحه تاني.
    """

    def _doc(self, **settings_):
        doc = self.inv.document
        doc["settings"]["intro_enabled"] = True
        doc["settings"].update(settings_)
        return B.normalize_document(doc)

    def _preview(self, doc):
        self.client.force_login(self.staff)
        return self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/preview/",
            data=json.dumps({"document": doc}), content_type="application/json",
        ).json()

    def test_preview_returns_the_intro_separately(self):
        data = self._preview(self._doc(intro_text="أهلاً بيكم"))
        self.assertIn("intro", data)
        self.assertIn("lb-intro", data["intro"])
        self.assertIn("أهلاً بيكم", data["intro"])
        # مش المفروض تتكرر جوّه html بتاع الأقسام
        self.assertNotIn("lb-intro", data["html"])

    def test_editing_intro_text_changes_the_preview(self):
        first = self._preview(self._doc(intro_text="نص أول"))["intro"]
        second = self._preview(self._doc(intro_text="نص تاني"))["intro"]
        self.assertIn("نص أول", first)
        self.assertIn("نص تاني", second)
        self.assertNotIn("نص أول", second)

    def test_editing_the_button_label_changes_the_preview(self):
        out = self._preview(self._doc(intro_button="ادخل يا حبيبي"))["intro"]
        self.assertIn("ادخل يا حبيبي", out)

    def test_disabling_the_intro_returns_empty(self):
        doc = self.inv.document
        doc["settings"]["intro_enabled"] = False
        self.assertEqual(self._preview(B.normalize_document(doc))["intro"], "")

    def test_intro_is_marked_editable_in_the_preview(self):
        out = self._preview(self._doc())["intro"]
        self.assertIn("data-intro-editable", out)

    def test_page_and_preview_render_the_same_intro_partial(self):
        """لو الاتنين ما استعملوش نفس الملف هيفرقوا مع الوقت."""
        self.inv.document = self._doc(intro_text="نص مشترك")
        self.inv.save(update_fields=["document"])
        page = self.client.get(self.inv.get_absolute_url()).content.decode()
        self.assertIn("نص مشترك", page)
        self.assertIn("نص مشترك", self._preview(self.inv.document)["intro"])


# ==========================================================================
class IntroNoteTests(BaseAppTest):
    """النوت بقى حقل يكتبه المصمّم بدل ما الافتتاحية تاخد اسم العميل."""

    def _intro(self, **settings_):
        """جزء الافتتاحية بس — ‎lb-kicker‎ موجود في الغلاف كمان، فمقارنة
        على الصفحة كلها بتكدب."""
        doc = self.inv.document
        doc["settings"]["intro_enabled"] = True
        doc["settings"].update(settings_)
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])
        body = self.client.get(self.inv.get_absolute_url()).content.decode()
        i = body.index('<div class="lb-intro')
        return body[i:body.index("lb-stage", i)]

    def test_the_intro_no_longer_prints_the_client_name(self):
        """كانت بتاخد name_one من تبويب البيانات غصب — دي بيانات العميل
        مش قرار تصميم، وكانت بتطلع في كل قالب."""
        self.assertNotIn(self.inv.name_one, self._intro())

    def test_nothing_shows_when_the_note_is_empty(self):
        self.assertNotIn("lb-kicker", self._intro())

    def test_the_note_shows_what_was_written(self):
        self.assertIn("محمد", self._intro(intro_note="محمد & فرح"))

    def test_the_note_is_escaped(self):
        body = self._intro(intro_note="<img src=x onerror=alert(1)>")
        self.assertNotIn("<img", body)

    def test_the_note_defaults_to_empty(self):
        field = next(f for f in B.editor_schema()["settings_fields"]
                     if f["key"] == "intro_note")
        self.assertEqual(field["default"], "")


# ==========================================================================
class IntroExitTests(BaseAppTest):
    """الافتتاحية بتقفل التمرير تحتها — لازم يفضل ليها مخرج دايماً."""

    def setUp(self):
        super().setUp()
        js = (Path(settings.BASE_DIR) / "static/js/invite.js").read_text("utf-8")
        i = js.index("function initIntro()")
        self.js = js[i:js.index("\n  }\n", i)]

    def _intro(self, **settings_):
        doc = self.inv.document
        doc["settings"]["intro_enabled"] = True
        doc["settings"].update(settings_)
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])
        return self.client.get(self.inv.get_absolute_url()).content.decode()

    def test_the_button_is_there_by_default(self):
        self.assertIn("data-intro-open", self._intro())

    def test_an_empty_label_removes_the_button(self):
        """مع فيديو الافتتاحية الزر زيادة: الضيف داس زر التشغيل خلاص."""
        self.assertNotIn("data-intro-open", self._intro(intro_button=""))

    def test_a_tap_anywhere_opens_it_when_there_is_no_button(self):
        seg = self.js[self.js.index('intro.addEventListener("click"'):]
        self.assertIn("if (btn)", seg[:400])
        self.assertIn("open();", seg[:400])

    def test_the_tap_does_not_skip_while_the_play_button_waits(self):
        """الضيف ساعتها ما شافش الافتتاحية أصلاً."""
        seg = self.js[self.js.index('intro.addEventListener("click"'):]
        self.assertIn("is-awaiting-play", seg[:220])

    def test_the_end_of_the_video_still_opens_it(self):
        self.assertIn('video.addEventListener("ended"', self.js)


# ==========================================================================
class IntroSoundTests(BaseAppTest):
    """صوت فيديو الافتتاحية — المتصفح بيمنع الصوت التلقائي، فالزر هو الحل."""

    def _render(self, **settings_):
        doc = self.inv.document
        doc["settings"]["intro_enabled"] = True
        doc["settings"].update(settings_)
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])
        return self.client.get(self.inv.get_absolute_url()).content.decode()

    def test_video_still_starts_muted(self):
        """مفيش متصفح بيسمح بالتشغيل التلقائي بصوت — لو شلنا muted الفيديو
        نفسه مش هيشتغل خالص."""
        body = self._render(intro_video="/media/x.mp4")
        self.assertIn("muted", body)
        self.assertIn("playsinline", body)

    def test_sound_button_appears_with_a_video(self):
        self.assertIn("data-intro-sound", self._render(intro_video="/media/x.mp4"))

    def test_sound_button_can_be_switched_off(self):
        body = self._render(intro_video="/media/x.mp4", intro_video_sound=False)
        self.assertNotIn("data-intro-sound", body)

    def test_no_sound_button_without_a_video(self):
        self.assertNotIn("data-intro-sound", self._render(intro_image="/media/x.webp"))

    def test_video_has_no_loop_so_it_can_open_on_end(self):
        """loop بيمنع حدث ended اللي بيفتح الدعوة لوحدها."""
        body = self._render(intro_video="/media/x.mp4")
        video_tag = body[body.index("<video"):body.index("</video>")]
        self.assertNotIn(" loop", video_tag)


# ==========================================================================
class TemplateHygieneTests(TestCase):
    """أخطاء قوالب بتعدّي من غير ما ترمي استثناء — بتظهر للمستخدم كنص."""

    def _templates(self):
        import os
        from django.conf import settings as dj
        root = dj.BASE_DIR / "templates"
        for dirpath, _dirs, files in os.walk(root):
            for name in files:
                if name.endswith(".html"):
                    path = os.path.join(dirpath, name)
                    yield path, open(path, encoding="utf-8").read()

    def test_no_multiline_hash_comments(self):
        """‎{# #}‎ في Django بيشتغل على سطر واحد بس.

        تعليق ممتد على سطرين مابيرميش خطأ — بيتطبع على الصفحة كنص عادي
        قدام المستخدم. حصل مرتين في المشروع ده، فبقى ليه اختبار.
        """
        bad = []
        for path, body in self._templates():
            for i, line in enumerate(body.splitlines(), 1):
                if "{#" in line and "#}" not in line.split("{#", 1)[1]:
                    bad.append(f"{path}:{i}")
        self.assertEqual(bad, [], "تعليق {# #} ممتد على أكتر من سطر: " + ", ".join(bad))

    def test_every_template_loads(self):
        """لو حد نسي {% load %} الصفحة بتقع وقت العرض مش وقت الفحص."""
        from django.template.loader import get_template
        from django.conf import settings as dj
        root = dj.BASE_DIR / "templates"
        for path, _ in self._templates():
            rel = str(path).replace(str(root) + "/", "")
            with self.subTest(template=rel):
                get_template(rel)


# ==========================================================================
class EmptyImportTests(TestCase):
    """قالب مستورد فاضي لازم يترفض بسبب واضح مش يتحفظ ويكتشفه المستخدم."""

    def _zip(self, files):
        import io as _io, zipfile
        buf = _io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            for n, d in files.items():
                z.writestr(n, d)
        return SimpleUploadedFile("t.zip", buf.getvalue(), "application/zip")

    SPA = ('<html><head><title>Site</title></head><body>'
           '<div id="__next"></div><script src="js/main.js"></script></body></html>')

    def test_javascript_built_page_is_allowed_with_a_local_script(self):
        document, _title, scripts, root_attrs = templateimport.build_document(
            *templateimport.parse_upload(self._zip({
                "index.html": self.SPA,
                "js/main.js": "document.body.dataset.loaded = 'yes';",
            })))
        self.assertGreaterEqual(len(document["blocks"]), 1)
        self.assertEqual(len(scripts), 1)
        self.assertTrue(scripts[0]["src"].startswith("/media/"))
        self.assertEqual(root_attrs, {"data-lb-spa": "true"})

    def test_nothing_is_saved_when_the_page_is_empty(self):
        before = Template.objects.count()
        with self.assertRaises(templateimport.ImportError_):
            templateimport.import_template(
                SimpleUploadedFile("a.html", self.SPA.encode(), "text/html"))
        self.assertEqual(Template.objects.count(), before)

    def test_loading_placeholder_is_still_treated_as_empty(self):
        """«جاري التحميل…» مش محتوى — ده نص مؤقت بيستبدله جافاسكربت."""
        page = self.SPA.replace('id="__next"></div>', 'id="__next">جاري التحميل…</div>')
        with self.assertRaises(templateimport.ImportError_):
            templateimport.build_document(
                *templateimport.parse_upload(self._zip({"index.html": page})))

    def test_a_real_page_still_imports(self):
        page = ("<html><head><title>عمر و ياسمين</title></head><body>"
                "<header><h1>عمر و ياسمين</h1>"
                "<p>يتشرفان بدعوتكم لحضور حفل زفافهما بقاعة الماسة</p></header>"
                "<section><h2>الموعد</h2>"
                "<p>الجمعة ٢٤ أكتوبر ٢٠٢٦ الساعة الثامنة مساءً</p></section>"
                "</body></html>")
        tpl = templateimport.import_template(
            SimpleUploadedFile("a.html", page.encode(), "text/html"))
        self.assertEqual(len(tpl.document["blocks"]), 2)

    def test_html_without_a_body_tag_still_imports(self):
        """قصاصات كتير مالهاش <body> — كانت بتترفض غلط."""
        page = ('<div class="hero"><h1>عمر و ياسمين</h1>'
                "<p>يتشرفان بدعوتكم لحضور حفل زفافهما في قاعة الماسة</p></div>")
        tpl = templateimport.import_template(
            SimpleUploadedFile("a.html", page.encode(), "text/html"))
        self.assertGreaterEqual(len(tpl.document["blocks"]), 1)


# ==========================================================================
class TemplateManageTests(BaseAppTest):
    """حذف/إخفاء القوالب من اللوحة — من غيرهم القالب الفاضي مالوش مخرج."""

    def _tpl(self, **kw):
        base = dict(name="مؤقت", slug="temp-x", category="classic",
                    source="import", document={"version": 1, "blocks": []})
        base.update(kw)
        return Template.objects.create(**base)

    def test_delete_removes_an_unused_template(self):
        tpl = self._tpl()
        self.client.force_login(self.staff)
        self.client.post("/dashboard/templates/",
                         {"action": "delete", "template": tpl.pk})
        self.assertFalse(Template.objects.filter(pk=tpl.pk).exists())

    def test_a_used_template_is_not_deleted(self):
        """الحذف هيسيب دعوات بغير قالب — بنمنعه ونقترح الإخفاء."""
        self.client.force_login(self.staff)
        res = self.client.post("/dashboard/templates/",
                               {"action": "delete", "template": self.template.pk},
                               follow=True)
        self.assertTrue(Template.objects.filter(pk=self.template.pk).exists())
        self.assertIn("اخفيه", res.content.decode())

    def test_toggle_hides_and_shows(self):
        tpl = self._tpl(is_active=True)
        self.client.force_login(self.staff)
        self.client.post("/dashboard/templates/",
                         {"action": "toggle", "template": tpl.pk})
        tpl.refresh_from_db()
        self.assertFalse(tpl.is_active)

    def test_delete_requires_staff(self):
        """غير مسجّل = تحويل لصفحة الدخول، والقالب مايتحذفش."""
        tpl = self._tpl()
        res = self.client.post("/dashboard/templates/",
                               {"action": "delete", "template": tpl.pk})
        self.assertIn("/login", res.get("Location", ""))
        self.assertTrue(Template.objects.filter(pk=tpl.pk).exists())

    def test_staff_can_open_template_editor(self):
        self.client.force_login(self.staff)
        res = self.client.get(f"/dashboard/templates/{self.template.pk}/editor/")
        self.assertEqual(res.status_code, 200)
        self.assertContains(res, "editor-document")
        self.assertContains(res, "تعديل قالب")
        self.assertNotContains(res, 'data-tab="data"')

    def test_staff_can_save_template_document(self):
        self.client.force_login(self.staff)
        document = self.template.get_document()
        document["theme"]["accent"] = "#123456"
        res = self.client.post(
            f"/dashboard/templates/{self.template.pk}/api/save/",
            data=json.dumps({"document": document, "fields": {}}),
            content_type="application/json",
        )
        self.assertEqual(res.status_code, 200)
        self.assertTrue(res.json()["ok"])
        self.template.refresh_from_db()
        self.assertEqual(self.template.get_document()["theme"]["accent"], "#123456")

    def test_template_demo_shows_latest_saved_document(self):
        # افتح النسخة القديمة أولاً لمحاكاة ضغط المستخدم على «معاينة» قبل التعديل.
        first = self.client.get(f"/templates/{self.template.slug}/preview/")
        self.assertEqual(first.status_code, 200)

        self.client.force_login(self.staff)
        document = self.template.get_document()
        document["theme"]["accent"] = "#123456"
        saved = self.client.post(
            f"/dashboard/templates/{self.template.pk}/api/save/",
            data=json.dumps({"document": document, "fields": {}}),
            content_type="application/json",
        )
        self.assertEqual(saved.status_code, 200)
        self.assertTrue(saved.json()["ok"])

        latest = self.client.get(f"/templates/{self.template.slug}/preview/")
        self.assertEqual(latest.status_code, 200)
        self.assertContains(latest, "#123456")

    def test_template_editor_requires_staff(self):
        user = User.objects.create_user("reader", password="Pass!12345x")
        self.client.force_login(user)
        res = self.client.get(f"/dashboard/templates/{self.template.pk}/editor/")
        self.assertEqual(res.status_code, 403)




# ==========================================================================
class FontLibraryTests(BaseAppTest):
    def test_staff_can_open_font_library(self):
        self.client.force_login(self.staff)
        response = self.client.get("/dashboard/fonts/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "مكتبة الخطوط")
        self.assertContains(response, "WOFF2")

    def test_staff_can_upload_font(self):
        self.client.force_login(self.staff)
        response = self.client.post(
            "/dashboard/fonts/",
            {
                "name": "خط الزفاف",
                "name_en": "Wedding Font",
                "family": "WeddingFont",
                "weight": "400",
                "style": "normal",
                "is_active": "on",
                "order": "0",
                "file": SimpleUploadedFile(
                    "wedding.woff2", b"fake-font", content_type="font/woff2"
                ),
            },
        )
        self.assertEqual(response.status_code, 302)
        font = CustomFont.objects.get(family="WeddingFont")
        self.assertEqual(font.name, "خط الزفاف")
        self.assertTrue(font.url)
        editor = self.client.get(f"/dashboard/invitations/{self.inv.pk}/editor/")
        self.assertContains(editor, "WeddingFont")


# ==========================================================================
class FavoriteLibraryTests(BaseAppTest):
    def test_staff_can_save_and_delete_favorite_block(self):
        self.client.force_login(self.staff)
        block = self.inv.get_document()["blocks"][0]
        response = self.client.post(
            "/dashboard/favorites/api/create/",
            data=json.dumps({"name": "إطار الزفاف", "block": block}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        favorite = FavoriteBlock.objects.get(name="إطار الزفاف")
        self.assertEqual(payload["favorite"]["blockType"], favorite.block_type)
        self.assertEqual(payload["favorite"]["block"]["type"], favorite.block_type)

        deleted = self.client.post(f"/dashboard/favorites/{favorite.pk}/delete/")
        self.assertEqual(deleted.status_code, 200)
        self.assertFalse(FavoriteBlock.objects.filter(pk=favorite.pk).exists())


# ==========================================================================
class ImportedAudioTests(TestCase):

    """قوالب الدعوات بتيجي ومعاها ملف موسيقى — ما نتجاهلوش."""

    def _zip(self, files):
        import io as _io, zipfile
        buf = _io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            for n, d in files.items():
                z.writestr(n, d)
        return SimpleUploadedFile("t.zip", buf.getvalue(), "application/zip")

    PAGE = ("<html><head><title>حفل</title></head><body>"
            "<h1>محمد و فرح</h1>"
            "<p>يتشرفان بدعوتكم لحضور حفل زفافهما في قاعة الياسمين</p>"
            "</body></html>")

    def test_audio_lands_in_the_music_library(self):
        before = MusicTrack.objects.count()
        templateimport.import_template(
            self._zip({"index.html": self.PAGE, "media/Music.mp3": b"ID3fake"}),
            name="قالب بموسيقى")
        self.assertEqual(MusicTrack.objects.count(), before + 1)
        track = MusicTrack.objects.latest("id")
        self.assertIn("قالب بموسيقى", track.name)
        self.assertTrue(track.url)

    def test_template_without_audio_adds_nothing(self):
        before = MusicTrack.objects.count()
        templateimport.import_template(self._zip({"index.html": self.PAGE}))
        self.assertEqual(MusicTrack.objects.count(), before)

    def test_audio_does_not_become_a_block(self):
        """الملف الصوتي أصل مستقل — مش قسم في الدعوة."""
        tpl = templateimport.import_template(
            self._zip({"index.html": self.PAGE, "media/x.mp3": b"ID3fake"}))
        self.assertNotIn(".mp3", json.dumps(tpl.document))


# ==========================================================================
class BlockLabelTests(TestCase):
    """اسم القسم في القائمة — من غيره القالب المستورد كله «كود HTML مخصص»."""

    def _import(self, body):
        page = f"<html><head><title>T</title></head><body>{body}</body></html>"
        return templateimport.import_template(
            SimpleUploadedFile("a.html", page.encode(), "text/html"))

    LONG = "<p>يتشرفان بدعوتكم لحضور حفل زفافهما في قاعة الياسمين بالقاهرة</p>"

    def test_heading_becomes_the_label(self):
        tpl = self._import(f"<section><h2>قصتنا</h2>{self.LONG}</section>")
        self.assertEqual(tpl.document["blocks"][0]["label"], "قصتنا")

    def test_entities_are_decoded_in_the_label(self):
        """«ليلى &amp; كريم» كان بيظهر بالكيان في القائمة."""
        tpl = self._import(f"<header><h1>ليلى &amp; كريم</h1>{self.LONG}</header>")
        self.assertEqual(tpl.document["blocks"][0]["label"], "ليلى & كريم")

    def test_known_class_names_map_to_arabic(self):
        tpl = self._import(f'<div class="preloader">{self.LONG}</div>')
        self.assertEqual(tpl.document["blocks"][0]["label"], "شاشة التحميل")

    def test_tag_name_is_the_next_fallback(self):
        tpl = self._import(f"<footer>{self.LONG}</footer>")
        self.assertEqual(tpl.document["blocks"][0]["label"], "الخاتمة")

    def test_label_survives_normalisation(self):
        doc = B.normalize_document({
            "blocks": [{"id": "b1", "type": "text", "label": "قسمي أنا"}],
        })
        self.assertEqual(doc["blocks"][0]["label"], "قسمي أنا")

    def test_label_is_trimmed_and_capped(self):
        doc = B.normalize_document({
            "blocks": [{"id": "b1", "type": "text", "label": "  " + "ط" * 200}],
        })
        self.assertEqual(len(doc["blocks"][0]["label"]), 60)

    def test_non_string_label_is_ignored(self):
        doc = B.normalize_document({
            "blocks": [{"id": "b1", "type": "text", "label": {"x": 1}}],
        })
        self.assertEqual(doc["blocks"][0]["label"], "")


# ==========================================================================
class ImportedEditingTests(TestCase):
    """تحرير القالب المستورد بصرياً — نص، موضع، لون — من غير ما تفتح الكود."""

    def _import(self, body, css=""):
        page = (f"<html><head><title>T</title>"
                f"{'<style>' + css + '</style>' if css else ''}</head>"
                f"<body>{body}</body></html>")
        return templateimport.import_template(
            SimpleUploadedFile("a.html", page.encode(), "text/html"))

    LONG = "<p>يتشرفان بدعوتكم لحضور حفل زفافهما في قاعة الياسمين بالقاهرة</p>"

    # ---- المواضع
    def test_el_slots_survive_normalisation(self):
        """el-N فيها شرطة، و_SLOT_RE مابيسمحش بيها — كانت بتتشال كلها."""
        doc = B.normalize_document({"blocks": [{
            "id": "imp-1", "type": "custom_html",
            "layout": {"el-3": {"dx": -15.5, "dy": -10.3}},
        }]})
        self.assertEqual(doc["blocks"][0]["layout"],
                         {"el-3": {"dx": -15.5, "dy": -10.3}})

    def test_el_slot_renders_positioning_css(self):
        doc = B.normalize_document({"blocks": [{
            "id": "imp-1", "type": "custom_html",
            "layout": {"el-3": {"dx": 4, "dy": -2}},
        }]})
        css = layout_css(doc["blocks"])
        self.assertIn('#imp-1 [data-move="el-3"]{--dx:4px;--dy:-2px}', css)

    def test_bogus_slot_names_are_still_refused(self):
        doc = B.normalize_document({"blocks": [{
            "id": "imp-1", "type": "custom_html",
            "layout": {"../etc": {"dx": 1}, "el-": {"dx": 1},
                       "el-12345678": {"dx": 1}, "EL-1": {"dx": 1}},
        }]})
        self.assertEqual(doc["blocks"][0]["layout"], {})

    # ---- data-move لازم ينجو من المنقّي
    def test_data_move_survives_sanitising(self):
        """من غيره العنصر بيفقد ارتباطه بموضعه المحفوظ أول ما تحفظ."""
        out = clean_html('<div data-move="el-7">نص</div>')
        self.assertIn('data-move="el-7"', out)

    def test_bad_data_move_value_is_dropped(self):
        out = clean_html('<div data-move="a b&quot;c">نص</div>')
        self.assertNotIn("data-move", out)

    def test_other_data_attributes_are_still_stripped(self):
        out = clean_html('<div data-track="spy" data-move="el-1">نص</div>')
        self.assertNotIn("data-track", out)
        self.assertIn("data-move", out)

    # ---- الكود مش في الواجهة
    def test_code_fields_sit_in_the_advanced_group(self):
        """الكود مش المفروض يكون أول حاجة يشوفها اللي رفع قالب جاهز."""
        spec = B.BLOCK_REGISTRY["custom_html"]
        for f in spec["props"]:
            if f["key"] in ("html", "css"):
                self.assertEqual(f["group"], "كود متقدّم")

    # ---- الألوان موجودة في الستايل عشان اللوحة تلاقيها
    def test_imported_css_keeps_its_colours_for_the_picker(self):
        tpl = self._import(f'<section class="s">{self.LONG}</section>',
                           css=".s{color:#b08948;background:#faf6ef}")
        stored = tpl.document["blocks"][0]["props"]["css"]
        self.assertIn("#b08948", stored)
        self.assertIn("#faf6ef", stored)


# ==========================================================================
class ElementStyleTests(TestCase):
    """لوحة العنصر بتكتب style مضمّن — لازم ينجو من المنقّي كله."""

    def test_font_and_size_survive(self):
        out = clean_html(
            '<h1 style="font-family:\'Great Vibes\', cursive;font-size:92px">ليلى</h1>')
        self.assertIn("font-family", out)
        self.assertIn("Great Vibes", out)
        self.assertIn("font-size:92px", out)

    def test_colour_weight_align_spacing_survive(self):
        out = clean_html('<p style="color:#b08948;background-color:#fff;'
                         'font-weight:700;text-align:center;letter-spacing:2px;'
                         'line-height:1.6">نص</p>')
        for want in ("color:#b08948", "background-color:#fff", "font-weight:700",
                     "text-align:center", "letter-spacing:2px", "line-height:1.6"):
            self.assertIn(want, out)

    def test_hiding_an_element_survives(self):
        self.assertIn("display:none", clean_html('<p style="display:none">نص</p>'))

    def test_dangerous_style_is_still_dropped(self):
        out = clean_html('<p style="font-size:20px;behavior:url(x.htc);'
                         'width:expression(alert(1))">نص</p>')
        self.assertIn("font-size:20px", out)
        self.assertNotIn("behavior", out)
        self.assertNotIn("expression", out)

    def test_external_url_in_style_is_still_dropped(self):
        """رابط خارجي في style بيسرّب زيارة الضيف لسيرفر تاني."""
        out = clean_html('<p style="background-image:url(https://evil.test/x.png)">ن</p>')
        self.assertNotIn("evil.test", out)

    def test_our_own_media_url_is_allowed(self):
        out = clean_html('<p style="background-image:url(/media/a/b.webp)">ن</p>')
        self.assertIn("/media/a/b.webp", out)


# ==========================================================================
class GuestPassTests(BaseAppTest):
    """تصريح الدخول: كود قصير + عدد دخلات + حالة نشط/مستخدم."""

    def _guest(self, **kw):
        base = dict(invitation=self.inv, name="كريم", entries_allowed=1)
        base.update(kw)
        return Guest.objects.create(**base)

    def test_pass_code_is_generated_and_readable(self):
        g = self._guest()
        self.assertTrue(g.pass_code.startswith("FRH-"))
        self.assertEqual(len(g.pass_code), 10)
        # الحروف اللي بتتلخبط في القراءة مستبعدة
        for ch in "IO01":
            self.assertNotIn(ch, g.pass_code[4:])

    def test_pass_codes_are_unique(self):
        codes = {self._guest(name=f"ض{i}").pass_code for i in range(40)}
        self.assertEqual(len(codes), 40)

    def test_entries_counting(self):
        g = self._guest(entries_allowed=3)
        self.assertEqual((g.entries_left, g.pass_status), (3, "active"))
        self.assertTrue(g.consume_entry())
        self.assertEqual((g.entries_used, g.entries_left), (1, 2))
        self.assertTrue(g.consume_entry())
        self.assertTrue(g.consume_entry())
        self.assertEqual(g.pass_status, "used")
        self.assertFalse(g.consume_entry())      # الرابعة مترفوضة
        self.assertEqual(g.entries_used, 3)

    def test_first_entry_sets_checked_in(self):
        g = self._guest()
        g.consume_entry()
        self.assertTrue(g.checked_in)
        self.assertIsNotNone(g.checked_in_at)

    def test_no_pass_when_allowance_is_zero(self):
        self.assertEqual(self._guest(entries_allowed=0).pass_status, "none")

    # ---- الصفحة
    def test_pass_page_shows_code_and_counts(self):
        g = self._guest(entries_allowed=3)
        g.consume_entry()
        body = self.client.get(
            f"/i/{self.inv.slug}/g/{g.token}/pass/").content.decode()
        self.assertIn(g.pass_code, body)
        self.assertIn("<svg", body)
        self.assertIn("تحميل الرمز صورة", body)

    def test_pass_page_needs_the_right_token(self):
        self._guest()
        self.assertEqual(
            self.client.get(f"/i/{self.inv.slug}/g/{'z' * 24}/pass/").status_code, 404)

    def test_qr_png_downloads(self):
        g = self._guest()
        res = self.client.get(f"/i/{self.inv.slug}/g/{g.token}/qr.png")
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["Content-Type"], "image/png")
        self.assertIn(g.pass_code, res["Content-Disposition"])
        self.assertTrue(res.content.startswith(b"\x89PNG"))


# ==========================================================================
class RsvpIssuesPassTests(BaseAppTest):
    """اللي بيأكد حضوره لازم يطلع بتصريح — حتى لو مش في كشف الضيوف."""

    def _rsvp(self, **kw):
        data = {"name": "منى", "status": "attending", "companions": 2}
        data.update(kw)
        return self.client.post(
            f"/i/{self.inv.slug}/rsvp/", data,
            HTTP_X_REQUESTED_WITH="XMLHttpRequest")

    def _enable_rsvp(self, max_companions=5):
        doc = self.inv.document
        for b in doc["blocks"]:
            if b["type"] == "rsvp":
                b["props"]["max_companions"] = max_companions
                break
        else:
            doc["blocks"].append({"id": "rsvp-1", "type": "rsvp",
                                  "props": {"max_companions": max_companions}})
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])

    def test_self_registered_guest_gets_a_pass(self):
        self._enable_rsvp()
        data = self._rsvp().json()
        self.assertTrue(data["ok"])
        self.assertIn("pass", data)
        g = Guest.objects.get(invitation=self.inv, name="منى")
        self.assertEqual(g.source, "rsvp")
        self.assertEqual(g.entries_allowed, 3)      # هي + اتنين
        self.assertEqual(data["pass"]["code"], g.pass_code)

    def test_pass_payload_has_qr_and_download_links(self):
        self._enable_rsvp()
        info = self._rsvp().json()["pass"]
        for key in ("code", "url", "qr", "download", "entries"):
            self.assertIn(key, info)
        self.assertIn("qr.png", info["download"])

    def test_declining_leaves_no_pass(self):
        self._enable_rsvp()
        data = self._rsvp(name="سامي", status="declined").json()
        self.assertNotIn("pass", data)
        self.assertFalse(Guest.objects.filter(name="سامي").exists())

    def test_named_guest_allowance_follows_companions(self):
        self._enable_rsvp()
        g = Guest.objects.create(invitation=self.inv, name="خالد",
                                 plus_ones_allowed=4, entries_allowed=1)
        self._rsvp(name="خالد", guest_token=g.token, companions=3)
        g.refresh_from_db()
        self.assertEqual(g.entries_allowed, 4)      # هو + تلاتة

    def test_allowance_never_drops_below_people_already_inside(self):
        """لو ٣ دخلوا خلاص، تعديل الرد لواحد مايخليش العدّاد سالب."""
        self._enable_rsvp()
        g = Guest.objects.create(invitation=self.inv, name="خالد",
                                 plus_ones_allowed=4, entries_allowed=4)
        for _ in range(3):
            g.consume_entry()
        self._rsvp(name="خالد", guest_token=g.token, companions=0)
        g.refresh_from_db()
        self.assertEqual(g.entries_allowed, 3)
        self.assertEqual(g.entries_left, 0)

    def test_changing_to_declined_cancels_an_unused_pass(self):
        self._enable_rsvp()
        g = Guest.objects.create(invitation=self.inv, name="خالد",
                                 plus_ones_allowed=2, entries_allowed=3)
        self._rsvp(name="خالد", guest_token=g.token, status="declined")
        g.refresh_from_db()
        self.assertEqual(g.entries_allowed, 0)
        self.assertEqual(g.pass_status, "none")


# ==========================================================================
class CheckinEntriesTests(BaseAppTest):
    """المسح بيستهلك دخلة — مش بيقلب مفتاح مرة واحدة."""

    def setUp(self):
        super().setUp()
        self.guest = Guest.objects.create(invitation=self.inv, name="كريم",
                                          entries_allowed=2)
        self.client.force_login(self.staff)

    def _scan(self, code=None):
        return self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/checkin/scan/",
            {"token": code or self.guest.token}).json()

    def test_each_scan_consumes_one_entry(self):
        first = self._scan()
        self.assertEqual((first["used"], first["left"], first["status"]),
                         (1, 1, "active"))
        second = self._scan()
        self.assertEqual((second["used"], second["left"], second["status"]),
                         (2, 0, "used"))

    def test_extra_scan_warns_instead_of_passing_silently(self):
        self._scan(); self._scan()
        third = self._scan()
        self.assertTrue(third["already"])
        self.assertIn("خلص", third["error"])
        self.guest.refresh_from_db()
        self.assertEqual(self.guest.entries_used, 2)   # ما زادش

    def test_short_code_works_when_the_qr_will_not_scan(self):
        out = self._scan(self.guest.pass_code)
        self.assertTrue(out["ok"])
        self.assertEqual(out["code"], self.guest.pass_code)

    def test_guest_without_allowance_is_refused(self):
        self.guest.grant_entries(0)
        out = self._scan()
        self.assertTrue(out["already"])
        self.assertIn("مالوش تصريح", out["error"])

    def test_totals_count_entries_not_rows(self):
        Guest.objects.create(invitation=self.inv, name="آخر", entries_allowed=3)
        out = self._scan()
        self.assertEqual(out["total"], 5)      # ٢ + ٣
        self.assertEqual(out["arrived"], 1)


# ==========================================================================
class GuestExportTests(BaseAppTest):
    """كشف الإكسل اللي بيتبعت للقاعة."""

    def setUp(self):
        super().setUp()
        self.client.force_login(self.staff)
        Guest.objects.create(invitation=self.inv, name="كريم عبد الله",
                             phone="01000000000", entries_allowed=3)
        Guest.objects.create(invitation=self.inv, name="منى", entries_allowed=1,
                             source="rsvp")

    def test_export_returns_a_workbook(self):
        res = self.client.get(f"/dashboard/invitations/{self.inv.pk}/guests/export.xlsx")
        self.assertEqual(res.status_code, 200)
        self.assertIn("spreadsheetml", res["Content-Type"])
        self.assertTrue(res.content.startswith(b"PK"))   # xlsx = أرشيف zip

    def test_sheet_has_a_row_per_guest_with_the_code(self):
        import io as _io
        from openpyxl import load_workbook
        res = self.client.get(f"/dashboard/invitations/{self.inv.pk}/guests/export.xlsx")
        wb = load_workbook(_io.BytesIO(res.content))
        ws = wb["الضيوف"]
        codes = [ws.cell(row=r, column=1).value for r in range(3, 5)]
        self.assertEqual(set(codes),
                         set(Guest.objects.values_list("pass_code", flat=True)))
        self.assertTrue(ws.sheet_view.rightToLeft)

    def test_sheet_embeds_a_qr_image_per_guest(self):
        """القاعة مالهاش نت على الباب — الرابط لوحده مايكفيش."""
        import io as _io
        from openpyxl import load_workbook
        res = self.client.get(f"/dashboard/invitations/{self.inv.pk}/guests/export.xlsx")
        wb = load_workbook(_io.BytesIO(res.content))
        self.assertEqual(len(wb["الضيوف"]._images), 2)

    def test_summary_sheet_totals_entries(self):
        import io as _io
        from openpyxl import load_workbook
        res = self.client.get(f"/dashboard/invitations/{self.inv.pk}/guests/export.xlsx")
        ws = load_workbook(_io.BytesIO(res.content))["ملخّص"]
        self.assertEqual(ws["B2"].value, 4)      # ٣ + ١ دخلات مسموحة

    def test_export_requires_staff(self):
        self.client.logout()
        res = self.client.get(f"/dashboard/invitations/{self.inv.pk}/guests/export.xlsx")
        self.assertNotEqual(res.status_code, 200)


# ==========================================================================
class IntroLibraryTests(BaseAppTest):
    """مكتبة الافتتاحيات — نفس فكرة مكتبة الموسيقى."""

    def setUp(self):
        super().setUp()
        self.clip = IntroVideo.objects.create(
            name="افتتاحية ذهبية", external_url="https://cdn.test/a.mp4", seconds=6)

    def test_page_requires_staff(self):
        self.assertNotEqual(self.client.get("/dashboard/intros/").status_code, 200)

    def test_page_lists_clips(self):
        self.client.force_login(self.staff)
        self.assertIn("افتتاحية ذهبية",
                      self.client.get("/dashboard/intros/").content.decode())

    def test_clip_needs_a_file_or_a_url(self):
        from system.forms import IntroVideoForm
        self.assertFalse(IntroVideoForm({"name": "بدون", "order": 0}).is_valid())
        self.assertTrue(IntroVideoForm(
            {"name": "برابط", "external_url": "https://x.test/a.mp4",
             "order": 0}).is_valid())

    def test_library_reaches_the_editor(self):
        self.client.force_login(self.staff)
        body = self.client.get(
            f"/dashboard/invitations/{self.inv.pk}/editor/").content.decode()
        self.assertIn("editor-intros", body)
        self.assertIn("cdn.test", body)

    def test_hidden_clips_do_not_reach_the_editor(self):
        IntroVideo.objects.update(is_active=False)
        self.client.force_login(self.staff)
        body = self.client.get(
            f"/dashboard/invitations/{self.inv.pk}/editor/").content.decode()
        self.assertIn('id="editor-intros" type="application/json">[]', body)

    def test_deleting_a_clip(self):
        self.client.force_login(self.staff)
        self.client.post("/dashboard/intros/",
                         {"action": "delete", "clip": self.clip.pk})
        self.assertFalse(IntroVideo.objects.filter(pk=self.clip.pk).exists())


# ==========================================================================
class IntroPlayButtonTests(BaseAppTest):
    """زر تشغيل الافتتاحية — لمسة الضيف بتسمح بالصوت من أول ثانية."""

    def _render(self, **settings_):
        doc = self.inv.document
        doc["settings"]["intro_enabled"] = True
        doc["settings"]["intro_video"] = "/media/x.mp4"
        doc["settings"].update(settings_)
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])
        return self.client.get(self.inv.get_absolute_url()).content.decode()

    def test_button_mode_shows_plain_text_without_effects(self):
        body = self._render(intro_play_mode="button", intro_play_label="افتح الدعوة")
        self.assertIn("data-intro-play", body)
        self.assertIn("data-intro-manual", body)
        self.assertIn("lb-intro-play--plain", body)
        tag = body[body.index("data-intro-play"):body.index("</button>", body.index("data-intro-play"))]
        self.assertIn("افتح الدعوة", tag)
        self.assertNotIn("<svg", tag)

    def test_intro_text_gear_settings_reach_preview(self):
        body = self._render(
            intro_text="النص التجريبي",
            intro_text_font="'Montserrat', sans-serif",
            intro_text_color="#f0c36a",
            intro_text_size=34,
        )
        text_start = body.index('class="lb-body"')
        text_tag = body[text_start:body.index(">", text_start)]
        self.assertIn("--intro-item-font:'Montserrat', sans-serif", text_tag)
        self.assertIn("--intro-item-color:#f0c36a", text_tag)
        self.assertIn("--intro-item-size:34px", text_tag)

    def test_old_intro_font_controls_are_hidden_from_editor_schema(self):
        fields = {f["key"]: f for f in B.editor_schema()["settings_fields"]}
        for key in (
            "intro_font", "intro_note_color", "intro_text_color",
            "intro_guest_name_color", "intro_button_color",
            "intro_play_color", "intro_play_bg_color",
        ):
            self.assertTrue(fields[key].get("editor_hidden"), key)

    def test_button_mode_does_not_autoplay(self):

        body = self._render(intro_play_mode="button")
        tag = body[body.index("<video"):body.index("</video>")]
        self.assertNotIn("autoplay", tag)
        # الوضع الافتراضي للصوت صامت، ويمكن تغييره من حقل الصوت.
        self.assertIn("muted", tag)

    def test_button_effects_mode_shows_a_play_button_with_effects(self):
        body = self._render(intro_play_mode="button_effects")
        self.assertIn("data-intro-play", body)
        self.assertIn("data-intro-manual", body)
        self.assertNotIn("lb-intro-play--no-effects", body)

    def test_auto_mode_is_still_muted_autoplay(self):
        tag_src = self._render(intro_play_mode="autoplay")
        tag = tag_src[tag_src.index("<video"):tag_src.index("</video>")]
        self.assertIn("autoplay", tag)
        self.assertIn("muted", tag)
        self.assertNotIn("data-intro-play", tag_src)

    def test_schema_exposes_only_the_three_start_modes(self):
        field = next(f for f in B.editor_schema()["settings_fields"]
                     if f["key"] == "intro_play_mode")
        self.assertEqual(field["default"], "autoplay")
        self.assertEqual(
            [option["value"] for option in field["options"]],
            ["autoplay", "button", "button_effects"],
        )
        self.assertNotIn(
            "intro_play_effects",
            {f["key"] for f in B.editor_schema()["settings_fields"]},
        )

    def test_legacy_effects_flag_maps_to_plain_button(self):
        doc = B.normalize_document({"settings": {"intro_play_effects": False}})
        self.assertEqual(doc["settings"]["intro_play_mode"], "button")
        self.assertFalse(doc["settings"]["intro_play_effects"])

    def test_legacy_autoplay_flag_maps_to_autoplay(self):
        doc = B.normalize_document({"settings": {"intro_autoplay": True}})
        self.assertEqual(doc["settings"]["intro_play_mode"], "autoplay")


# ==========================================================================
class FullBleedWidthTests(BaseAppTest):
    """«ملء الشاشة» — القسم بيكسر حدود كارت الدعوة على أي مقاس جهاز."""

    def setUp(self):
        super().setUp()
        # قسم الفيديو مربوط بميزة في الباقة — من غيرها بيتخفي من الصفحة
        self.plan.features = list(self.plan.features) + ["video"]
        self.plan.save(update_fields=["features"])

    def _render(self, width):
        doc = B.normalize_document({"blocks": [
            {"type": "video", "id": "vid-1",
             "props": {"url": "/media/clip.mp4"},
             "style": {"width": width}},
        ]})
        self.inv.document = doc
        self.inv.save(update_fields=["document"])
        return self.client.get(self.inv.get_absolute_url()).content.decode()

    def test_screen_width_is_an_offered_choice(self):
        values = [o["value"] for o in B.WIDTH_CHOICES]
        self.assertIn("screen", values)
        # «كامل العرض» لسه موجود — ده عرض الكارت مش الشاشة
        self.assertIn("full", values)

    def test_screen_width_emits_its_class(self):
        self.assertIn("lb--w-screen", self._render("screen"))

    def test_full_is_not_screen(self):
        body = self._render("full")
        self.assertIn("lb--w-full", body)
        self.assertNotIn("lb--w-screen", body)

    def test_bogus_width_falls_back_to_normal(self):
        self.assertIn("lb--w-normal", self._render("100vw"))

    def _rule(self):
        css = (Path(settings.BASE_DIR) / "static/css/invite.css").read_text("utf-8")
        i = css.index(".lb.lb--w-screen {")
        return css[i:css.index("}", i)]

    def test_breakout_happens_on_the_section_not_the_inner(self):
        """‎.lb‎ عليه overflow:hidden — لو الخروج من ‎.lb-inner‎ يتقص."""
        rule = self._rule()
        self.assertIn("100vw", rule)
        self.assertIn("calc(50% - 50vw)", rule)
        self.assertNotIn(".lb-inner", rule)

    def test_selector_outranks_the_mobile_padding_rule(self):
        """‎@media(max-width:640px){.lb{padding-inline:18px}}‎ بيجي بعده في
        الملف، وبنفس الأولوية كان بيكسب — فالفيديو كان بيفضل ٣٥٤ من ٣٩٠."""
        css = (Path(settings.BASE_DIR) / "static/css/invite.css").read_text("utf-8")
        self.assertIn(".lb.lb--w-screen {", css)
        # الكلاس المكرّر لازم ييجي قبل قاعدة الموبايل عشان يبقى المقصود
        self.assertIn("padding-inline: 0", self._rule())


# ==========================================================================
class VideoBlockSourceTests(BaseAppTest):
    """قسم الفيديو — رفع ملف مش لينك بس، ونسبة عرض/ارتفاع."""

    def setUp(self):
        super().setUp()
        # قسم الفيديو مربوط بميزة في الباقة — من غيرها بيتخفي من الصفحة
        self.plan.features = list(self.plan.features) + ["video"]
        self.plan.save(update_fields=["features"])

    def _spec(self, key):
        block = B.editor_schema()["blocks"]["video"]
        return next(f for f in block["props"] if f["key"] == key)

    def test_url_field_offers_upload_not_just_a_link(self):
        spec = self._spec("url")
        self.assertEqual(spec["type"], "media")
        self.assertEqual(spec["media_kind"], "video")

    def _render(self, props):
        doc = B.normalize_document({"blocks": [
            {"type": "video", "id": "vid-1", "props": props},
        ]})
        self.inv.document = doc
        self.inv.save(update_fields=["document"])
        return self.client.get(self.inv.get_absolute_url()).content.decode()

    def test_uploaded_file_path_survives_normalisation(self):
        body = self._render({"url": "/media/assets/2026/08/clip.webm"})
        self.assertIn('data-video="/media/assets/2026/08/clip.webm"', body)

    def test_youtube_link_still_works(self):
        body = self._render({"url": "https://youtu.be/abc12345"})
        self.assertIn("https://youtu.be/abc12345", body)

    def test_aspect_default_is_16x9(self):
        self.assertIn("lb-video--16x9", self._render({"url": "/media/c.mp4"}))

    def test_aspect_choice_is_rendered(self):
        self.assertIn("lb-video--9x16",
                      self._render({"url": "/media/c.mp4", "aspect": "9x16"}))

    def test_bogus_aspect_falls_back(self):
        body = self._render({"url": "/media/c.mp4", "aspect": "gzip; rm -rf"})
        self.assertIn("lb-video--16x9", body)
        self.assertNotIn("rm -rf", body)

    def test_controls_are_on_by_default(self):
        """الدعوات الموجودة ماينفعش يختفي منها الشريط فجأة."""
        self.assertTrue(self._spec("controls")["default"])
        self.assertNotIn("data-no-controls", self._render({"url": "/media/c.mp4"}))

    def test_controls_can_be_switched_off(self):
        self.assertIn("data-no-controls",
                      self._render({"url": "/media/c.mp4", "controls": False}))

    def test_sound_is_off_by_default(self):
        """فيديو بيطلع صوت من غير ما حد يطلبه بيخضّ الضيف."""
        self.assertFalse(self._spec("sound")["default"])
        self.assertNotIn("data-sound", self._render({"url": "/media/c.mp4"}))

    def test_sound_flag_reaches_the_markup(self):
        self.assertIn("data-sound", self._render(
            {"url": "/media/c.mp4", "autoplay": True, "sound": True}))

    def test_media_field_still_blocks_javascript_scheme(self):
        """الحقل كان type=url وليه فلتر. تحويله لـmedia ماينفعش يشيله."""
        doc = B.normalize_document({"blocks": [
            {"type": "video", "id": "v-1",
             "props": {"url": "javascript:alert(1)"}},
        ]})
        self.assertEqual(doc["blocks"][0]["props"]["url"], "")

    def test_image_fields_block_it_too(self):
        doc = B.normalize_document({"blocks": [
            {"type": "video", "id": "v-1",
             "props": {"url": "/x.mp4", "poster": "vbscript:msgbox 1"}},
        ]})
        self.assertEqual(doc["blocks"][0]["props"]["poster"], "")

    def test_inline_image_data_url_is_kept(self):
        """المحرر بيولّد data:image عند القص — مانكسرهاش."""
        tiny = ("data:image/png;base64,"
                "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAAC0lEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg==")
        doc = B.normalize_document({"blocks": [
            {"type": "video", "id": "v-1", "props": {"url": "/x.mp4", "poster": tiny}},
        ]})
        self.assertEqual(doc["blocks"][0]["props"]["poster"], tiny)

    def test_non_image_data_url_is_dropped(self):
        doc = B.normalize_document({"blocks": [
            {"type": "video", "id": "v-1",
             "props": {"url": "data:text/html;base64,PHNjcmlwdD4="}},
        ]})
        self.assertEqual(doc["blocks"][0]["props"]["url"], "")


# ==========================================================================
class VideoUploadLimitTests(BaseAppTest):
    """حد الفيديو أعلى من حد الصورة — مقطع فرح حقيقي بيعدّي ٨ ميجا."""

    def _post(self, raw, name, ctype):
        self.client.force_login(self.staff)
        return self.client.post(
            f"/dashboard/invitations/{self.inv.pk}/api/upload/",
            {"file": SimpleUploadedFile(name, raw, ctype)},
        )

    def test_video_over_the_image_limit_is_accepted(self):
        """قبل التعديل ده كان أي مقطع أكبر من ٨ ميجا بيترفض."""
        import shutil, subprocess, tempfile, os
        if not shutil.which("ffmpeg"):
            self.skipTest("ffmpeg غير متاح")
        path = tempfile.mktemp(suffix=".mp4")
        # مقطع ضوضاء مش بينضغط — بيطلع أكبر من ٨ ميجا فعلاً
        subprocess.run(
            ["ffmpeg", "-y", "-loglevel", "error", "-f", "lavfi",
             "-i", "nullsrc=size=640x360:rate=25:duration=2",
             "-vf", "geq=random(1)*255:128:128", "-c:v", "libx264",
             "-preset", "ultrafast", "-qp", "0", "-pix_fmt", "yuv420p", path],
            check=True, timeout=180)
        raw = open(path, "rb").read()
        os.unlink(path)
        if not (8 * 1024 * 1024 < len(raw) < video.MAX_UPLOAD_BYTES):
            self.skipTest(f"المقطع المولَّد خرج بره النطاق المطلوب: {len(raw)}")
        res = self._post(raw, "clip.mp4", "video/mp4").json()
        self.assertTrue(res["ok"], res)

    def test_video_over_its_own_limit_is_refused(self):
        raw = b"\0" * (video.MAX_UPLOAD_BYTES + 1)
        res = self._post(raw, "big.mp4", "video/mp4")
        self.assertEqual(res.status_code, 400)
        self.assertIn("٤٠", res.json()["error"])

    def test_image_limit_did_not_change(self):
        raw = b"\0" * (8 * 1024 * 1024 + 1)
        res = self._post(raw, "big.png", "image/png")
        self.assertEqual(res.status_code, 400)
        self.assertIn("٨", res.json()["error"])

    def test_intro_library_uses_the_same_limit(self):
        """حدّين مختلفين لنفس الملف حسب مكان الرفع = ارتباك."""
        from system.forms import IntroVideoForm
        form = IntroVideoForm(
            {"name": "x", "order": 0},
            {"file": SimpleUploadedFile("a.mp4", b"\0" * (9 * 1024 * 1024), "video/mp4")},
        )
        form.is_valid()
        self.assertNotIn("file", form.errors)


# ==========================================================================
class IntroPlayLabelTests(BaseAppTest):
    """نص اختياري على زر تشغيل الافتتاحية."""

    def _render(self, label=None):
        doc = self.inv.document
        doc["settings"]["intro_enabled"] = True
        doc["settings"]["intro_video"] = "/media/x.mp4"
        doc["settings"]["intro_play_mode"] = "button"

        if label is not None:
            doc["settings"]["intro_play_label"] = label
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])
        return self.client.get(self.inv.get_absolute_url()).content.decode()

    def _button(self, body):
        i = body.index("data-intro-play")
        start = body.rindex("<button", 0, i)
        return body[start:body.index("</button>", start)]

    def test_default_is_icon_only(self):
        btn = self._button(self._render())
        self.assertNotIn("lb-intro-play--label", btn)
        self.assertNotIn("lb-intro-play-text", btn)
        self.assertIn('aria-label="تشغيل الفيديو"', btn)

    def test_label_appears_inside_the_button(self):
        btn = self._button(self._render("اضغط لتشغيل الفيديو"))
        self.assertIn("lb-intro-play--label", btn)
        self.assertIn("اضغط لتشغيل الفيديو", btn)
        # والأيقونة لسه موجودة جنب الكلام
        self.assertIn("<svg", btn)

    def test_label_replaces_the_generic_aria_label(self):
        """قارئ الشاشة ماينفعش يقول حاجتين مختلفتين عن نفس الزر."""
        btn = self._button(self._render("ابدأ العرض"))
        self.assertIn('aria-label="ابدأ العرض"', btn)
        self.assertNotIn('aria-label="تشغيل الفيديو"', btn)

    def test_label_is_escaped(self):
        btn = self._button(self._render('<img src=x onerror=alert(1)>'))
        self.assertNotIn("<img", btn)
        self.assertIn("&lt;img", btn)

    def test_label_is_ignored_in_auto_mode(self):
        doc = self.inv.document
        doc["settings"].update({"intro_enabled": True, "intro_video": "/media/x.mp4",
                                                                "intro_play_mode": "autoplay",

                                "intro_play_label": "ابدأ"})
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])
        body = self.client.get(self.inv.get_absolute_url()).content.decode()
        self.assertNotIn("data-intro-play", body)

    def test_field_default_is_empty(self):
        field = next(f for f in B.editor_schema()["settings_fields"]
                     if f["key"] == "intro_play_label")
        self.assertEqual(field["default"], "")


# ==========================================================================
class VideoPlayerPathTests(TestCase):
    """الملف المرفوع بيتركّب على طول، والطرف التالت بضغطة."""

    def setUp(self):
        self.js = (Path(settings.BASE_DIR) / "static/js/invite.js").read_text("utf-8")
        i = self.js.index("function initVideo()")
        self.body = self.js[i:self.js.index("\n  }\n", i)]

    def test_local_file_does_not_wait_for_a_click(self):
        """كانت ضغطة زيادة على ملف مالوش طرف تالت — ومع «زي ما هو»
        كانت بتسيب شريط رفيع مكان الفيديو لحد ما يتضغط."""
        self.assertIn("if (!yt && !vimeo) {", self.body)
        guard = self.body.index("if (!yt && !vimeo) {")
        click = self.body.index("lb-video-play")
        self.assertLess(guard, click, "لازم المسار المحلي ييجي قبل زر الضغط")

    def test_third_party_still_needs_a_click(self):
        self.assertIn("lb-video-play", self.body)
        self.assertIn("youtube-nocookie", self.js)

    def test_natural_ratio_is_published_to_the_container(self):
        self.assertIn("loadedmetadata", self.body)
        self.assertIn("--vid-ratio", self.body)

    def test_css_consumes_that_ratio_with_a_fallback(self):
        css = (Path(settings.BASE_DIR) / "static/css/invite.css").read_text("utf-8")
        # المطلوب: الـCSS بيستهلك المتغيّر اللي الـJS بيكتبه وليه قيمة
        # احتياطية — مش شكل المسافات نفسه
        self.assertRegex(css, r"var\(--vid-ratio,\s*16\s*/\s*9\)")

    def test_autoplay_without_a_gesture_is_still_muted(self):
        """من غير لمسة من الضيف المتصفح بيمنع الصوت — والصامت هو الطريق
        الوحيد اللي الفيديو بيشتغل بيه أصلاً."""
        seg = self.body[self.body.index("if (autoplay && !withSound)"):]
        self.assertIn("v.muted = true", seg[:120])

    def test_sound_is_gated_on_the_intro_screen(self):
        """الصوت مربوط بلمسة الضيف على زر فتح الدعوة — مفيش إذن تاني."""
        self.assertIn('doc.querySelector(".lb-intro")', self.body)
        self.assertIn("lb:intro-open", self.body)

    def test_the_gesture_is_spent_inside_its_own_listener(self):
        """صلاحية اللمسة لحظة الضغطة بس. لو استنّينا الفيديو يوصل
        للشاشة الإذن بيبقى راح والصوت بيترفض."""
        seg = self.body[self.body.index('"lb:intro-open"'):]
        self.assertIn("v.play()", seg[:400])

    def test_refused_sound_falls_back_to_muted(self):
        """الافتتاحية ممكن تتفتح بالعدّاد التلقائي — يبقى مفيش لمسة."""
        seg = self.body[self.body.index('"lb:intro-open"'):]
        self.assertIn("v.muted = true", seg[:700])

    def test_music_is_ducked_while_a_loud_video_plays(self):
        """موسيقى وفيديو صوتي فوق بعض = دوشة."""
        self.assertIn("__lbMusic", self.body)
        seg = self.body[self.body.index("var duck ="):]
        self.assertIn("m.pause()", seg[:250])

    def test_hidden_controls_come_back_when_autoplay_is_blocked(self):
        """وضع توفير الطاقة في iOS بيرفض التشغيل التلقائي. من غير الرجوع
        ده الضيف بيبص على صورة ساكنة من غير أي زر يشغّلها بيه."""
        seg = self.body[self.body.index("var tryPlay"):]
        self.assertIn("catch", seg[:300])
        self.assertIn("v.controls = true", seg[:300])

    def test_autoplay_waits_for_the_section_to_be_on_screen(self):
        """كل فيديوهات الدعوة تشتغل مع بعض = موبايل بيهنّج وداتا بتتحرق."""
        self.assertIn("IntersectionObserver", self.body)
        self.assertIn("v.pause()", self.body)

    def test_guests_cannot_download_the_video_from_the_controls(self):
        self.assertIn("nodownload", self.body)

    def test_a_bare_video_is_clickable(self):
        """من غير شريط ومن غير تشغيل تلقائي الفيديو صورة ساكنة —
        فالضغطة عليه لازم تبقى هي زر التشغيل."""
        seg = self.body[self.body.index("} else if (noControls) {"):]
        self.assertIn('addEventListener("click"', seg[:400])


# ==========================================================================
class VideoHeightCapTests(TestCase):
    """فيديو طولي بعرض الشاشة كان بيطلع ٢٥٦٠px ارتفاع على ديسك توب."""

    def setUp(self):
        css = (Path(settings.BASE_DIR) / "static/css/invite.css").read_text("utf-8")
        i = css.index(".lb-video {")
        self.rule = css[i:css.index("}", i)]
        self.css = css

    def test_there_is_a_height_cap(self):
        self.assertIn("max-width: calc(", self.rule)
        self.assertIn("vh", self.rule)

    def test_cap_is_on_width_so_the_ratio_survives(self):
        """لو السقف كان على ‎max-height‎ النسبة كانت هتتكسر والفيديو
        يتحط جوّه شرايط سودا بدل ما ياخد مقاسه."""
        self.assertNotIn("max-height", self.rule)

    def test_ratio_and_cap_read_the_same_variable(self):
        """لو كل واحد ليه مصدر مختلف بيفترقوا عند أول تعديل."""
        self.assertIn("aspect-ratio: var(--ar", self.rule)
        self.assertIn("var(--ar", self.rule[self.rule.index("max-width"):])

    def test_every_aspect_choice_sets_that_variable(self):
        for key in ("16x9", "4x3", "1x1", "9x16", "auto"):
            with self.subTest(key=key):
                i = self.css.index(f".lb-video--{key} ")
                self.assertIn("--ar:", self.css[i:self.css.index("}", i)])

    def test_choices_in_python_match_the_classes_in_css(self):
        spec = next(f for f in B.editor_schema()["blocks"]["video"]["props"]
                    if f["key"] == "aspect")
        for o in spec["options"]:
            with self.subTest(o=o["value"]):
                self.assertIn(f".lb-video--{o['value']} ", self.css)


# ==========================================================================
class AutoScrollTests(BaseAppTest):
    """التمرير التلقائي — مفتاح تشغيل، وسرعة، ويقف أول ما الضيف يلمس."""

    def _render(self, **s):
        doc = self.inv.document
        doc["settings"].update(s)
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])
        return self.client.get(self.inv.get_absolute_url()).content.decode()

    def test_default_is_off(self):
        """ميزة بتحرّك الصفحة تحت إيد الواحد ماينفعش تبقى الافتراضي."""
        field = next(f for f in B.editor_schema()["settings_fields"]
                     if f["key"] == "auto_scroll")
        self.assertIs(field["default"], False)
        body = self._render()
        self.assertIn('"enabled": false', body)

    def test_enabled_ships_its_settings(self):
        body = self._render(auto_scroll=True, auto_scroll_speed="fast",
                            auto_scroll_delay=5, auto_scroll_loop=True)
        node = body[body.index('id="invite-scroll"'):]
        node = node[:node.index("</script>")]
        self.assertIn('"enabled": true', node)
        self.assertIn('"speed": "fast"', node)
        self.assertIn('"delay": 5', node)
        self.assertIn('"loop": true', node)

    def test_bogus_speed_falls_back(self):
        body = self._render(auto_scroll=True, auto_scroll_speed="\"};alert(1)//")
        self.assertIn('"speed": "normal"', body)
        self.assertNotIn("alert(1)", body)

    def test_it_is_off_inside_the_editor(self):
        """الصفحة ماينفعش تفضل نازلة لوحدها وإنت بتعدّل فيها."""
        from system.views import _scroll_config
        on = {"auto_scroll": True, "auto_scroll_speed": "fast"}
        self.assertTrue(_scroll_config(on)["enabled"])
        self.assertFalse(_scroll_config(on, editable=True)["enabled"])

    def test_delay_is_clamped(self):
        body = self._render(auto_scroll=True, auto_scroll_delay=9999)
        self.assertIn('"delay": 20', body)      # الحد الأعلى في الحقل

    # -- ضمانات في ملف الجافاسكربت ---------------------------------------
    def _js(self):
        js = (Path(settings.BASE_DIR) / "static/js/invite.js").read_text("utf-8")
        i = js.index("function initAutoScroll()")
        return js[i:js.index("\n  // -----", i)]

    def test_reduced_motion_disables_it_completely(self):
        """الإعداد ده طلب صريح من ناس الحركة بتتعبهم."""
        self.assertIn("prefers-reduced-motion", self._js())

    def test_user_interaction_stops_it(self):
        body = self._js()
        for ev in ("wheel", "touchstart", "pointerdown", "keydown"):
            with self.subTest(ev=ev):
                self.assertIn(f'"{ev}"', body)
        self.assertIn("stoppedByUser", body)

    def test_the_control_button_is_exempt_from_that(self):
        """من غير الاستثناء ده الضغط على «تشغيل» كان بيوقّفه في نفس اللحظة."""
        self.assertIn("btn.contains(e.target)", self._js())

    def test_it_waits_for_the_intro_to_close(self):
        """الافتتاحية بتقفل التمرير أصلاً، فمالوش معنى يبدأ قبلها."""
        self.assertIn("lb:intro-open", self._js())
        js = (Path(settings.BASE_DIR) / "static/js/invite.js").read_text("utf-8")
        self.assertIn('dispatchEvent(new CustomEvent("lb:intro-open"))', js)


# ==========================================================================
class PreviewCtaTests(BaseAppTest):
    """شريط «عجبك القالب؟» — في معاينة القالب بس."""

    def setUp(self):
        super().setUp()
        self.cfg = SiteSetting.load()
        self.cfg.whatsapp_number = "+20 155-940 3203"
        self.cfg.facebook_url = "https://facebook.com/farha"
        self.cfg.save()

    def _preview(self):
        return self.client.get(f"/templates/{self.template.slug}/preview/").content.decode()

    def test_bar_shows_with_the_template_name(self):
        body = self._preview()
        self.assertIn("data-preview-cta", body)
        self.assertIn("عجبك القالب ده؟", body)
        self.assertIn(self.template.name, body)

    def test_whatsapp_link_carries_the_template_name(self):
        body = self._preview()
        self.assertIn("wa.me/201559403203", body)
        # الاسم متشفّر في الرابط
        self.assertIn(quote(self.template.name), body)

    def test_number_is_normalised_for_wa_me(self):
        """wa.me بيرفض + والمسافات والشرط — بيدّي صفحة غلط من غير شرح."""
        self.assertEqual(SiteSetting.load().whatsapp_digits, "201559403203")

    def test_facebook_link_is_there(self):
        self.assertIn("facebook.com/farha", self._preview())

    def test_toggle_hides_the_whole_bar(self):
        self.cfg.preview_cta_enabled = False
        self.cfg.save()
        self.assertNotIn("data-preview-cta", self._preview())

    def test_each_icon_has_its_own_toggle(self):
        self.cfg.whatsapp_enabled = False
        self.cfg.save()
        body = self._preview()
        self.assertNotIn("wa.me", body)
        self.assertIn("facebook.com/farha", body)   # فيسبوك لسه ظاهر

    def test_bar_disappears_when_nothing_is_left_to_show(self):
        """شريط فيه نص وبس ومفيش أي زرار = ضوضاء."""
        self.cfg.whatsapp_enabled = False
        self.cfg.facebook_enabled = False
        self.cfg.save()
        self.assertNotIn("data-preview-cta", self._preview())

    def test_it_never_shows_on_a_real_invitation(self):
        """دعوة العميل مش مكان إعلان — الشريط للمعاينة بس."""
        body = self.client.get(self.inv.get_absolute_url()).content.decode()
        self.assertNotIn("data-preview-cta", body)

    def test_message_placeholder_is_replaced(self):
        self.cfg.whatsapp_message = "عايز {template} لو سمحت"
        self.cfg.save()
        body = self._preview()
        self.assertIn(quote(f"عايز {self.template.name} لو سمحت"), body)
        self.assertNotIn("{template}", body)

    def test_settings_are_one_row_only(self):
        """سجلين إعدادات = مين اللي شغّال؟"""
        SiteSetting.objects.create(preview_cta_text="تاني")
        self.assertEqual(SiteSetting.objects.count(), 1)
        self.assertEqual(SiteSetting.load().preview_cta_text, "تاني")


# ==========================================================================
class PreviewCtaOnAFreshInstallTests(BaseAppTest):
    """أول ما تنصّب المشروع الشريط مفعّل بس الرقم فاضي.

    فبيختفي بالكامل، واللوحة كانت بتقول «ظاهر»، ومفيش أي رسالة تفسّر —
    وقعد الواحد يدوّر في الكود على حاجة موجودة أصلاً وشغّالة صح.
    """

    def _preview(self):
        return self.client.get(
            f"/templates/{self.template.slug}/preview/").content.decode()

    def _dash(self):
        self.client.force_login(self.staff)
        return self.client.get("/dashboard/site/").content.decode()

    def test_the_defaults_leave_it_enabled_but_contactless(self):
        cfg = SiteSetting.load()
        self.assertTrue(cfg.preview_cta_enabled)
        self.assertEqual(cfg.whatsapp_number, "")
        self.assertEqual(cfg.facebook_url, "")

    def test_so_the_bar_does_not_render(self):
        self.assertNotIn("data-preview-cta", self._preview())

    def test_the_dashboard_says_why_instead_of_staying_silent(self):
        self.assertIn("data-cta-warning", self._dash())

    def test_the_status_line_does_not_claim_it_is_visible(self):
        dash = self._dash()
        self.assertIn("مفعّل بس مخفي", dash)

    def test_one_number_is_enough_to_bring_it_back(self):
        cfg = SiteSetting.load()
        cfg.whatsapp_number = "+20 100 000 0000"
        cfg.save()
        self.assertIn("data-preview-cta", self._preview())
        self.assertNotIn("data-cta-warning", self._dash())

    def test_facebook_alone_is_enough_too(self):
        cfg = SiteSetting.load()
        cfg.facebook_url = "https://facebook.com/farha"
        cfg.save()
        self.assertIn("data-preview-cta", self._preview())
        self.assertNotIn("data-cta-warning", self._dash())

    def test_the_page_and_the_dashboard_read_one_rule(self):
        """لما الشرط كان مكرر في الاتنين، اللوحة كذبت على المستخدم."""
        views = (Path(settings.BASE_DIR) / "system/views.py").read_text("utf-8")
        i = views.index("def _preview_cta(")
        self.assertIn("cfg.preview_cta_ready", views[i:i + 900])

    def test_a_disabled_toggle_is_not_reported_as_a_problem(self):
        cfg = SiteSetting.load()
        cfg.preview_cta_enabled = False
        cfg.save()
        self.assertNotIn("data-cta-warning", self._dash())


# ==========================================================================
class OrdersToggleTests(BaseAppTest):
    """قفل نموذج «اطلب دعوتك» من اللوحة بدل ما يتعمل كومنت في الكود.

    الكومنت كان هيخفي القسم بس ويسيب حاجتين مكسورين: الـPOST على ‎/‎
    لسه شغّال وأي حد يقدر يبعت طلب، وأزرار «اطلب دعوتك» في النڤ
    والقوالب والباقات بتودّي على مرساة ‎#order‎ مش موجودة.
    """

    def setUp(self):
        super().setUp()
        self.cfg = SiteSetting.load()
        self.cfg.whatsapp_number = "+201000000000"
        self.cfg.save()

    def _off(self):
        self.cfg.orders_enabled = False
        self.cfg.save()

    def _home(self):
        return self.client.get("/").content.decode()

    def test_open_by_default_so_nothing_changes_for_existing_sites(self):
        self.assertTrue(SiteSetting.load().orders_enabled)
        self.assertIn("إرسال الطلب", self._home())

    def test_closing_it_removes_the_form(self):
        self._off()
        self.assertNotIn("إرسال الطلب", self._home())

    def test_closing_it_also_refuses_a_direct_post(self):
        """إخفاء القسم مش قفل — المسار لسه موجود."""
        self._off()
        r = self.client.post("/", {
            "name": "زائر", "phone": "01000000000", "plan": self.plan.pk,
            "template": self.template.pk, "event_type": "زفاف",
        })
        self.assertEqual(r.status_code, 403)
        self.assertEqual(Order.objects.count(), 0)

    def test_an_open_form_still_takes_orders(self):
        r = self.client.post("/", {
            "name": "زائر", "phone": "01000000000", "plan": self.plan.pk,
            "template": self.template.pk, "event_type": "زفاف",
        })
        self.assertIn(r.status_code, (200, 302))

    def test_the_buttons_point_at_whatsapp_instead_of_a_dead_anchor(self):
        self._off()
        body = self._home()
        self.assertNotIn('href="#order"', body)
        self.assertIn("wa.me/201000000000", body)

    def test_with_no_number_the_buttons_disappear_rather_than_break(self):
        self.cfg.whatsapp_number = ""
        self.cfg.orders_enabled = False
        self.cfg.save()
        body = self._home()
        self.assertNotIn('href="#order"', body)
        self.assertNotIn("اطلب على واتساب", body)


# ==========================================================================
class PlanAddonShowcaseTests(BaseAppTest):
    """الإضافات بقت عرض سعر تحت كل باقة — مش اختيار في نموذج."""

    def setUp(self):
        super().setUp()
        self.video = PlanAddon.objects.create(name="اضافه فيديو", price=200)
        self.song = PlanAddon.objects.create(name="اضافه اغنيه", price=300)
        self.song.plans.add(self.basic)          # مربوطة بباقة واحدة بس

    def _home(self):
        return self.client.get("/").content.decode()

    def test_the_price_is_shown_under_the_plan(self):
        body = self._home()
        self.assertIn("إضافات اختيارية", body)
        self.assertIn("اضافه فيديو", body)
        self.assertIn("+200", body)

    def test_an_addon_with_no_plans_shows_under_all_of_them(self):
        i = self._home().count("اضافه فيديو")
        self.assertGreaterEqual(i, Plan.objects.filter(is_active=True).count())

    def test_an_addon_tied_to_one_plan_shows_only_there(self):
        body = self._home()
        section = body[body.index('id="plans"'):body.index('id="order"')]
        self.assertEqual(section.count("اضافه اغنيه"), 1)

    def test_it_is_a_price_list_not_a_picker(self):
        """عرض بس — مفيش شيك بوكس ولا مجموع في قسم الباقات."""
        body = self._home()
        section = body[body.index('id="plans"'):body.index('id="order"')]
        self.assertNotIn("checkbox", section)
        self.assertNotIn("data-addons-total", section)

    def test_an_inactive_addon_is_not_advertised(self):
        self.video.is_active = False
        self.video.save()
        self.assertNotIn("اضافه فيديو", self._home())


# ==========================================================================
class WhatsAppFloatTests(BaseAppTest):
    """أيقونة واتساب عائمة في صفحات الموقع، رقمها من لوحة التحكم."""

    def setUp(self):
        super().setUp()
        self.cfg = SiteSetting.load()
        self.cfg.whatsapp_number = "+20 155-940 3203"
        self.cfg.save()

    def test_it_shows_on_the_home_page(self):
        body = self.client.get("/").content.decode()
        self.assertIn("data-wa-float", body)
        self.assertIn("wa.me/201559403203", body)

    def test_it_shows_on_the_gallery_too(self):
        self.assertIn("data-wa-float",
                      self.client.get("/templates/").content.decode())

    def test_the_ready_message_rides_along(self):
        self.cfg.whatsapp_cta_message = "عايز أستفسر"
        self.cfg.save()
        self.assertIn(quote("عايز أستفسر"), self.client.get("/").content.decode())

    def test_no_number_means_no_button(self):
        self.cfg.whatsapp_number = ""
        self.cfg.save()
        self.assertNotIn("data-wa-float", self.client.get("/").content.decode())

    def test_its_own_toggle_turns_it_off(self):
        self.cfg.whatsapp_float_enabled = False
        self.cfg.save()
        self.assertNotIn("data-wa-float", self.client.get("/").content.decode())

    def test_the_preview_cta_toggle_does_not_control_it(self):
        """التنين وسايل تواصل مختلفة — قفل واحد مايقفلش التاني."""
        self.cfg.preview_cta_enabled = False
        self.cfg.save()
        self.assertIn("data-wa-float", self.client.get("/").content.decode())

    def test_it_never_shows_on_a_guest_invitation(self):
        """دعوة العميل مش مكان إعلان عننا."""
        body = self.client.get(self.inv.get_absolute_url()).content.decode()
        self.assertNotIn("data-wa-float", body)

    def test_it_never_shows_in_the_dashboard(self):
        self.client.force_login(self.staff)
        self.assertNotIn("data-wa-float",
                         self.client.get("/dashboard/").content.decode())

    def test_the_number_is_editable_from_the_dashboard(self):
        self.client.force_login(self.staff)
        body = self.client.get("/dashboard/site/").content.decode()
        for name in ("whatsapp_number", "whatsapp_float_enabled",
                     "whatsapp_cta_message", "orders_enabled"):
            with self.subTest(name=name):
                self.assertIn(f'name="{name}"', body)


# ==========================================================================
class PlanAddonTests(BaseAppTest):
    """إضافات بسعر زيادة فوق الباقة."""

    def setUp(self):
        super().setUp()
        self.music = PlanAddon.objects.create(
            name="موسيقى بالخلفية", code="music", price=50)
        self.fast = PlanAddon.objects.create(
            name="تسليم ٢٤ ساعة", code="", price=100)

    def test_addon_page_lists_them(self):
        self.client.force_login(self.staff)
        body = self.client.get("/dashboard/plans/").content.decode()
        self.assertIn("موسيقى بالخلفية", body)
        self.assertIn("تسليم ٢٤ ساعة", body)

    def test_page_is_staff_only(self):
        self.assertEqual(self.client.get("/dashboard/plans/").status_code, 302)
        self.client.force_login(self.normal)
        self.assertEqual(self.client.get("/dashboard/plans/").status_code, 403)

    def test_create_edit_toggle_delete(self):
        self.client.force_login(self.staff)
        self.client.post("/dashboard/plans/", {
            "name": "معرض صور", "code": "gallery", "price": "75", "sort_order": "0",
            "is_active": "on",
        })
        new = PlanAddon.objects.get(name="معرض صور")
        self.assertEqual(new.code, "gallery")

        self.client.post("/dashboard/plans/", {
            "addon": new.pk, "name": "معرض صور موسّع", "code": "gallery",
            "price": "90", "sort_order": "0", "is_active": "on",
        })
        new.refresh_from_db()
        self.assertEqual(new.name, "معرض صور موسّع")
        self.assertEqual(int(new.price), 90)

        self.client.post("/dashboard/plans/", {"addon": new.pk, "action": "toggle"})
        new.refresh_from_db()
        self.assertFalse(new.is_active)

        self.client.post("/dashboard/plans/", {"addon": new.pk, "action": "delete"})
        self.assertFalse(PlanAddon.objects.filter(pk=new.pk).exists())

    def test_unknown_feature_code_is_refused(self):
        """إضافة بمفتاح مش موجود بتتباع ومابتفتحش حاجة — والعميل مش
        هيعرف إن مفيش فرق."""
        form = PlanAddonForm({"name": "x", "code": "not_a_feature",
                              "price": "10", "sort_order": "0"})
        self.assertFalse(form.is_valid())
        self.assertIn("code", form.errors)

    def test_blank_code_is_fine(self):
        """الإضافة ممكن تبقى خدمة مش ميزة — زي التسليم السريع."""
        form = PlanAddonForm({"name": "تسليم سريع", "code": "",
                              "price": "100", "sort_order": "0"})
        self.assertTrue(form.is_valid(), form.errors)

    def test_purchased_addon_is_hidden_not_deleted(self):
        """مسحها كان هيمسح تاريخ الطلب معاها."""
        order = self._order(self.music)
        self.client.force_login(self.staff)
        self.client.post("/dashboard/plans/", {"addon": self.music.pk, "action": "delete"})
        self.music.refresh_from_db()
        self.assertFalse(self.music.is_active)
        self.assertTrue(order.order_addons.exists())

    # -- الطلب -------------------------------------------------------------
    def _order(self, *addons, plan=None):
        form = OrderForm({
            "customer_name": "يوسف", "customer_phone": "0100000001",
            "plan": (plan or self.plan).pk, "event_type": "زفاف",
            "names": "يوسف و سارة",
            "addons": [a.pk for a in addons],
        })
        self.assertTrue(form.is_valid(), form.errors)
        return form.save()

    def test_order_total_adds_the_addons(self):
        order = self._order(self.music, self.fast)
        self.assertEqual(order.addons_total, Decimal("150"))
        self.assertEqual(order.total_price, self.plan.price + Decimal("150"))

    def test_price_at_purchase_is_frozen(self):
        """تغيير سعر الإضافة بعدين ماينفعش يعيد كتابة طلب قديم."""
        order = self._order(self.music)
        self.music.price = Decimal("500")
        self.music.save(update_fields=["price"])
        order.refresh_from_db()
        self.assertEqual(order.addons_total, Decimal("50"))

    def test_name_at_purchase_is_frozen_too(self):
        order = self._order(self.music)
        self.music.name = "اسم تاني خالص"
        self.music.save(update_fields=["name"])
        self.assertEqual(order.order_addons.first().name, "موسيقى بالخلفية")

    def test_bought_feature_opens_on_the_invitation(self):
        """اللي دفع في «موسيقى» فوق باقة مافيهاش موسيقى لازم ياخدها."""
        bare = Plan.objects.create(name="بسيطة", slug="bare", price=300,
                                   features=["countdown"])
        inv = Invitation.objects.create(
            customer=self.customer, template=self.template, plan=bare,
            name_one="أ", name_two="ب", status="published",
            event_date=timezone.now() + timedelta(days=10),
            document=self.template.get_document(),
        )
        self.assertNotIn("music", inv.allowed_features)

        order = self._order(self.music, plan=bare)
        order.invitation = inv
        order.save(update_fields=["invitation"])
        inv.refresh_from_db()
        self.assertIn("music", inv.allowed_features)
        self.assertIn("countdown", inv.allowed_features)   # ومزايا الباقة باقية

    def test_service_addon_opens_nothing(self):
        order = self._order(self.fast)
        self.assertEqual(order.addon_features, set())

    def test_addon_limited_to_other_plans_is_refused(self):
        """الإخفاء في المتصفح مش حماية — الفورم بيتبعت من غير جافاسكربت."""
        self.music.plans.set([self.basic])
        form = OrderForm({
            "customer_name": "يوسف", "customer_phone": "0100000002",
            "plan": self.plan.pk, "event_type": "زفاف",
            "addons": [self.music.pk],
        })
        self.assertFalse(form.is_valid())
        self.assertIn("addons", form.errors)

    def test_unrestricted_addon_works_with_any_plan(self):
        self.assertTrue(self.music.available_for(self.plan))
        self.assertTrue(self.music.available_for(self.basic))

    def test_addons_appear_in_the_order_form_on_the_site(self):
        body = self.client.get("/").content.decode()
        self.assertIn("موسيقى بالخلفية", body)
        self.assertIn("data-addons", body)

    def test_inactive_addon_is_not_offered(self):
        self.music.is_active = False
        self.music.save(update_fields=["is_active"])
        self.assertNotIn("موسيقى بالخلفية", self.client.get("/").content.decode())

    def test_orders_table_shows_the_addons_and_total(self):
        self._order(self.music)
        self.client.force_login(self.staff)
        body = self.client.get("/dashboard/orders/").content.decode()
        self.assertIn("موسيقى بالخلفية", body)
        self.assertIn(str(int(self.plan.price + 50)), body)


# ==========================================================================
class PlanQuickEditTests(BaseAppTest):
    """تعديل اسم وسعر الباقة من الجدول من غير ما تفتح صفحة تانية."""

    def setUp(self):
        super().setUp()
        self.client.force_login(self.staff)

    def _post(self, **extra):
        data = {"action": "plan", "plan": self.plan.pk, "name": self.plan.name,
                "price": str(self.plan.price)}
        data.update(extra)
        return self.client.post("/dashboard/plans/", data)

    def test_price_and_name_are_saved(self):
        self._post(name="مميزة جداً", price="1250", is_active="on")
        self.plan.refresh_from_db()
        self.assertEqual(self.plan.name, "مميزة جداً")
        self.assertEqual(self.plan.price, Decimal("1250"))
        self.assertTrue(self.plan.is_active)

    def test_unchecked_box_actually_hides_the_plan(self):
        """مربّع اختيار مش متعلّم مابيتبعتش أصلاً — لازم نتعامل مع غيابه."""
        self._post()                       # من غير is_active
        self.plan.refresh_from_db()
        self.assertFalse(self.plan.is_active)

    def test_negative_price_is_refused(self):
        old = self.plan.price
        self._post(price="-100")
        self.plan.refresh_from_db()
        self.assertEqual(self.plan.price, old)

    def test_garbage_price_does_not_crash(self):
        old = self.plan.price
        res = self._post(price="مية جنيه")
        self.assertEqual(res.status_code, 302)
        self.plan.refresh_from_db()
        self.assertEqual(self.plan.price, old)

    def test_blank_old_price_clears_it(self):
        self.plan.old_price = Decimal("2000")
        self.plan.save(update_fields=["old_price"])
        self._post(old_price="", is_active="on")
        self.plan.refresh_from_db()
        self.assertIsNone(self.plan.old_price)

    def test_features_are_not_touched(self):
        """التعديل السريع للسعر ماينفعش يمسح مزايا الباقة بالغلط."""
        before = list(self.plan.features)
        self._post(price="900", is_active="on")
        self.plan.refresh_from_db()
        self.assertEqual(self.plan.features, before)

    def test_the_row_forms_live_outside_the_table(self):
        """‎<form>‎ ابن مباشر لـ‎<tr>‎ مش HTML صالح والمتصفح بينقله."""
        body = self.client.get("/dashboard/plans/").content.decode()
        table = body[body.index("<h2 style=\"font-size:18px;margin-top:28px\">الباقات"):]
        table = table[:table.index("</table>")]
        head, rows = table.split("<tbody>", 1)
        self.assertIn('id="plan-', head)          # الفورمات قبل الجدول
        self.assertNotIn("<form", rows)           # ومفيش فورم جوّه الصفوف


# ==========================================================================
class CssVariableTests(TestCase):
    """كل ‎var(--x)‎ **من غير قيمة احتياطية** لازم يكون ليه مصدر.

    الحكاية دي جت من غلطة حقيقية: كتبت ‎var(--border)‎ في ‎site.css‎
    والاسم ده موجود في ‎invite.css‎ بس. النتيجة إن السطر كله بيبقى غير
    صالح — الكارت طلع من غير حدود ومن غير إبراز للمختار، **من غير أي
    رسالة خطأ في أي مكان**. ده أسوأ نوع غلط: بيسكت.

    ‎var(--x, 62px)‎ بقيمة احتياطية مالهاش المشكلة دي — بتنزل على
    الاحتياطي وخلاص، وده استخدام مقصود مش سهو.
    """

    # الاسم، وبعده فاصلة لو فيه قيمة احتياطية
    VAR = re.compile(r"var\(\s*(--[a-z0-9-]+)\s*(,?)")
    DEF = re.compile(r"(--[a-z0-9-]+)\s*:")

    def _read(self, rel):
        return (Path(settings.BASE_DIR) / rel).read_text("utf-8")

    def _required(self, css):
        """المتغيرات المستخدمة من غير قيمة احتياطية."""
        return {name for name, comma in self.VAR.findall(css) if not comma}

    def test_site_css_defines_everything_it_uses(self):
        css = self._read("static/css/site.css")
        missing = sorted(self._required(css) - set(self.DEF.findall(css)))
        self.assertEqual(missing, [], f"متغيرات مستخدمة ومش متعرّفة: {missing}")

    def test_invite_css_variables_come_from_somewhere(self):
        """‎invite.css‎ بياخد متغيرات كتير من المحرك وقت العرض، فاللي
        مش متعرّف في الملف لازم يكون بيتكتب من بايثون أو الجافاسكربت."""
        css = self._read("static/css/invite.css")
        missing = self._required(css) - set(self.DEF.findall(css))
        parts = [
            self._read("system/renderer.py"),
            self._read("system/templatetags/invite.py"),
            self._read("static/js/invite.js"),
            self._read("static/js/editor.js"),
        ]
        for folder in ("templates/blocks", "templates/invitations"):
            parts += [p.read_text("utf-8")
                      for p in (Path(settings.BASE_DIR) / folder).glob("*.html")]
        sources = "\n".join(parts)
        orphans = sorted(v for v in missing if v not in sources)
        self.assertEqual(orphans, [], f"متغيرات مالهاش مصدر خالص: {orphans}")

    def test_the_check_would_have_caught_the_real_bug(self):
        """ضمان إن الفحص نفسه شغّال مش بيعدّي كل حاجة."""
        fake = ".x { border: 1px solid var(--does-not-exist); }"
        self.assertEqual(self._required(fake), {"--does-not-exist"})
        # ومع قيمة احتياطية مايتبلّغش عنه
        self.assertEqual(self._required(".x { color: var(--nope, red); }"), set())


# ==========================================================================
class ElementInspectorTests(TestCase):
    """لوحة «العنصر المحدَّد» جوّه القسم المستورد.

    كانت بتعرض نفس أدوات النص مهما كان العنصر — يعني لو حدّدت صورة
    تلاقي قدامك خط وحجم خط ومحاذاة وتباعد حروف، وتحرّكهم ومايحصلش
    حاجة. أداة مالهاش تأثير أسوأ من أداة ناقصة.
    """

    def setUp(self):
        js = (Path(settings.BASE_DIR) / "static/js/editor.js").read_text("utf-8")
        i = js.index("function buildElementGroup(")
        self.body = js[i:js.index("\n  function ", i + 10)]

    def _guard_span(self, start_at):
        """مدى أول حارس ‎if (!isImg)‎ بعد ‎start_at‎ — بعدّ الأقواس مش بالشكل."""
        at = self.body.index("if (!isImg && !isMap) {", start_at)
        opening = self.body.index("{", at)
        depth = 0
        for j in range(opening, len(self.body)):
            if self.body[j] == "{":
                depth += 1
            elif self.body[j] == "}":
                depth -= 1
                if depth == 0:
                    return opening, j
        raise AssertionError("حارس مفتوح ومش مقفول")

    def _guarded(self, probe):
        """هل الحقل ده واقع جوّه حارس ‎if (!isImg)‎؟"""
        at = self.body.index(probe)
        pos = 0
        while True:
            try:
                start, end = self._guard_span(pos)
            except ValueError:
                return False
            if start < at < end:
                return True
            pos = end

    def test_an_image_is_recognised_by_its_tag(self):
        """الصورة ممكن تكون هي العنصر نفسه أو جوّه العنصر المحدَّد."""
        self.assertIn('var imageNode = tag === "IMG" ? node : node.querySelector("img")',
                      self.body)
        self.assertIn("var isImg = !!imageNode;", self.body)

    def test_the_font_controls_are_hidden_for_images(self):
        for probe in ('"font-family"', '"font-size"', '"font-weight"', '"text-align"'):
            with self.subTest(probe=probe):
                self.assertTrue(self._guarded(probe), f"{probe} مش متحجوب عن الصورة")

    def test_the_spacing_controls_are_hidden_for_images(self):
        for probe in ('"letter-spacing"', '"line-height"'):
            with self.subTest(probe=probe):
                self.assertTrue(self._guarded(probe), f"{probe} مش متحجوب عن الصورة")

    def test_text_colour_goes_but_background_stays(self):
        """‎color‎ مالوش أي أثر على ‎<img>‎، لكن الخلفية بتبان ورا PNG شفاف."""
        self.assertIn('isImg ? [["background-color"', self.body)

    def test_an_image_gets_a_crop_button(self):
        seg = self.body[self.body.index("if (isImg) {"):]
        self.assertIn("قصّ الصورة", seg)
        self.assertIn("openCropper(", seg)

    def test_crop_refuses_an_image_that_is_not_in_the_library(self):
        """القص بيشتغل على الأصل المحفوظ — صورة جاية مع قالب مستورد
        مالهاش أصل عندنا، فالزر لازم يقول كده مش يفشل بصمت."""
        seg = self.body[self.body.index("قصّ الصورة"):]
        self.assertIn("ASSETS", seg[:900])
        self.assertIn("if (!asset)", seg[:900])

    def test_an_image_gets_controls_that_actually_do_something(self):
        seg = self.body[self.body.index("if (isImg) {"):]
        self.assertIn('"width"', seg)
        self.assertIn('"border-radius"', seg)

    def test_the_guard_helper_would_catch_a_regression(self):
        """ضمان إن الفحص نفسه بيفرّق — مش بيرجع True على طول."""
        self.assertFalse(self._guarded("var isImg = !!imageNode;"))

    def test_editor_only_classes_are_not_shown_as_the_element_name(self):
        """‎lb-el-picked‎ كلاس بيحطه المحرر وقت الاختيار، مش كلاس القالب —
        كان بيطلع في اسم العنصر («صورة · lb-el-picked»)."""
        self.assertIn('"lb-el-picked": 1', self.body)


# ==========================================================================
class ImageSwapDragTests(TestCase):
    """سحب صورة وإفلاتها فوق صورة تانية = تبديل مش تحريك."""

    def setUp(self):
        js = (Path(settings.BASE_DIR) / "static/js/editor.js").read_text("utf-8")
        self.js = js
        i = js.index("function swapCandidate(")
        self.pick = js[i:js.index("\n  function ", i + 10)]

    def test_the_target_must_be_another_image(self):
        self.assertIn('t.tagName !== "IMG"', self.pick)
        self.assertIn("t === node", self.pick)

    def test_the_target_must_be_in_the_same_block(self):
        """الحفظ بيسلسل قسم واحد بس — تبديل بين قسمين كان هيضيّع نص التبديل."""
        self.assertIn('t.closest("[data-block]") !== holder', self.pick)

    def test_the_edges_do_not_count(self):
        """الصور في شبكة جنب بعض — زحزحة صغيرة كانت هتبقى تبديل بالغلط."""
        self.assertIn("SWAP_INNER", self.js)
        self.assertIn("r.left + mx", self.pick)

    def test_the_target_is_highlighted_before_release(self):
        self.assertIn("lb-el-swap", self.js)
        css = (Path(settings.BASE_DIR) / "static/css/invite.css").read_text("utf-8")
        self.assertIn(".lb-el-swap", css)

    def test_swapping_moves_srcset_too(self):
        """‎srcset‎ بتكسب على ‎src‎ — لو سبناها الصورة القديمة تفضل ظاهرة."""
        i = self.js.index("function swapImages(")
        seg = self.js[i:self.js.index("\n  function ", i + 10)]
        self.assertIn("srcset", seg)

    def test_the_nudge_is_undone_so_both_stay_in_place(self):
        i = self.js.index("if (!cancel && swapWith)")
        seg = self.js[i:i + 400]
        self.assertIn("applySlotOffset(node, d.dx0, d.dy0)", seg)
        self.assertIn("snapshot()", seg)


# ==========================================================================
class CropUsesTheOriginalTests(BaseAppTest):
    """نافذة القص كانت بتعرض النسخة المقصوصة وتقص من الأصل.

    ‎api_crop‎ بيقص من ‎asset.source‎ (الأصل قبل أي قص) عشان ما يحصلش فقد
    جودة متراكم. لكن الـJSON اللي بيروح للمحرر ما كانش فيه ‎source‎
    خالص، فالنافذة كانت بتعرض ‎asset.url‎ — النسخة المقصوصة. يعني الكادر
    اللي بترسمه على صورة، بيتطبّق على صورة تانية.
    """

    def _asset(self):
        from django.core.files.base import ContentFile
        from system.models import Asset
        return Asset.objects.create(
            file=ContentFile(b"x", name="a.webp"),
            source=ContentFile(b"y", name="a-source.webp"),
            kind="image", original_name="a.webp", invitation=self.inv,
        )

    def test_the_model_can_point_at_the_original(self):
        a = self._asset()
        self.assertNotEqual(a.source_url, a.url)

    def test_the_editor_payload_carries_it(self):
        a = self._asset()
        self.client.force_login(self.staff)
        body = self.client.get(
            f"/dashboard/invitations/{self.inv.pk}/editor/").content.decode()
        i = body.index('id="editor-assets"')
        blob = body[i:body.index("</script>", i)]
        self.assertIn("source", blob)
        self.assertIn(a.source.name.rsplit("/", 1)[-1], blob)

    def test_the_assets_api_carries_it(self):
        a = self._asset()
        self.client.force_login(self.staff)
        data = self.client.get(
            f"/dashboard/invitations/{self.inv.pk}/api/assets/").json()
        row = next(r for r in data["assets"] if r["id"] == a.pk)
        self.assertEqual(row["source"], a.source_url)

    def test_the_cropper_prefers_the_original(self):
        js = (Path(settings.BASE_DIR) / "static/js/editor.js").read_text("utf-8")
        i = js.index("function openCropper(")
        self.assertIn("asset.source || asset.url", js[i:i + 300])


# ==========================================================================
class DocumentWallpaperTests(BaseAppTest):
    """وول بيبر على الدعوة كلها — الخلفية لقسم واحد موجودة أصلاً."""

    def _vars(self, **theme):
        doc = self.inv.document
        doc["theme"].update(theme)
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])
        body = self.client.get(self.inv.get_absolute_url()).content.decode()
        i = body.index('<body class="lb-doc"')
        return body[i:body.index(">", i)]

    def test_no_image_means_none_not_a_broken_url(self):
        self.assertIn("--doc-bg:none", self._vars())

    def test_the_image_reaches_the_page(self):
        self.assertIn("/media/wall.webp", self._vars(bg_image="/media/wall.webp"))

    def test_a_semicolon_in_the_url_cannot_break_the_other_variables(self):
        """كل المتغيّرات في سمة ‎style‎ واحدة مفصولة بـ‎;‎."""
        out = self._vars(bg_image="/media/a;b.webp")
        self.assertNotIn("/media/a;b.webp", out)
        self.assertIn("%3B", out)
        self.assertIn("--doc-bg-attach", out)      # اللي بعده لسه سليم

    def test_no_overlay_means_no_veil_layer(self):
        self.assertIn("--doc-bg-veil:none", self._vars(bg_image="/media/w.webp"))

    def test_the_overlay_becomes_a_gradient_layer(self):
        out = self._vars(bg_image="/media/w.webp", bg_overlay=40)
        self.assertIn("rgba(0,0,0,0.4)", out)

    def test_fixed_is_off_by_default(self):
        self.assertIn("--doc-bg-attach:scroll", self._vars())
        self.assertIn("--doc-bg-attach:fixed", self._vars(bg_fixed=True))

    def test_the_css_consumes_both_layers_in_order(self):
        css = (Path(settings.BASE_DIR) / "static/css/invite.css").read_text("utf-8")
        i = css.index(".lb-stage {")
        rule = css[i:css.index("}", i)]
        # التعتيم لازم ييجي **قبل** الصورة — أول طبقة هي الأعلى في CSS
        self.assertIn("var(--doc-bg-veil, none), var(--doc-bg, none)", rule)
        self.assertIn("background-color: var(--bg)", rule)

    def test_a_section_can_still_have_its_own_background(self):
        """الطلب كان «للدعوة كلها ولا لأقسام معينة» — التانية موجودة أصلاً."""
        keys = [f["key"] for f in B.editor_schema()["blocks"]["text"]["style"]]
        self.assertIn("bg_image", keys)


# ==========================================================================
class DocumentTranslationTests(BaseAppTest):
    """النسخة الإنجليزية — نصوص يكتبها المصمّم بإيده، مفيش ترجمة آلية."""

    def _doc(self, table=None):
        doc = self.inv.document
        doc["i18n"] = {"en": table or {}}
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])
        return self.inv.document

    def _page(self, lang=""):
        url = self.inv.get_absolute_url() + (f"?lang={lang}" if lang else "")
        return self.client.get(url).content.decode()

    def _hero_key(self, field="kicker"):
        hero = next(b for b in self.inv.document["blocks"] if b["type"] == "hero")
        return f"{hero['id']}.{field}"

    # ---- التخزين
    def test_a_fresh_document_has_an_empty_table(self):
        self.assertEqual(B.normalize_document({})["i18n"], {"en": {}})

    def test_junk_is_thrown_away(self):
        doc = B.normalize_document({"i18n": {"en": {
            "ok": "Fine", "bad": 12, "": "x", "blank": "   ",
        }}})
        self.assertEqual(doc["i18n"]["en"], {"ok": "Fine"})

    def test_a_huge_value_is_capped(self):
        doc = B.normalize_document({"i18n": {"en": {"k": "a" * 9000}}})
        self.assertEqual(len(doc["i18n"]["en"]["k"]), B.MAX_I18N_VALUE)

    # ---- زرار اللغة
    def test_no_translation_means_no_button(self):
        """دعوة من غير ترجمة مالهاش لازمة تعرض زرار بيرجّع نفس الكلام."""
        self._doc()
        self.assertNotIn("data-lang-toggle", self._page())

    def test_one_line_is_enough_to_show_it(self):
        self._doc({self._hero_key(): "Wedding invitation"})
        self.assertIn("data-lang-toggle", self._page())

    def test_asking_for_english_without_a_translation_stays_arabic(self):
        self._doc()
        body = self._page("en")
        self.assertIn('<html lang="ar"', body)

    # ---- التبديل
    def test_the_english_text_replaces_the_arabic(self):
        self._doc({self._hero_key(): "Wedding invitation"})
        body = self._page("en")
        self.assertIn("Wedding invitation", body)
        self.assertNotIn("دعوة زفاف", body)

    def test_an_untranslated_field_stays_arabic(self):
        """نص ناقص أحسن من فراغ في وش الضيف."""
        self._doc({self._hero_key(): "Wedding invitation"})
        body = self._page("en")
        self.assertIn("Wedding invitation", body)
        self.assertIn("يتشرفان بدعوتكم", body)      # السطر الفرعي لسه عربي

    def test_arabic_is_untouched_by_default(self):
        self._doc({self._hero_key(): "Wedding invitation"})
        body = self._page()
        self.assertIn("دعوة زفاف", body)
        self.assertNotIn("Wedding invitation", body)

    def test_the_page_flips_to_ltr(self):
        self._doc({self._hero_key(): "Wedding invitation"})
        body = self._page("en")
        self.assertIn('<html lang="en"', body)
        self.assertIn('dir="ltr"', body)

    # ---- بيانات المناسبة
    def test_the_names_come_from_the_table_too(self):
        """الأسماء عايشة في جدول الدعوة مش المستند، وبرضو بيشوفها الضيف."""
        self._doc({"data.name_one": "Layla", "data.venue": "Jasmine Hall"})
        body = self._page("en")
        self.assertIn("Layla", body)
        self.assertIn("Jasmine Hall", body)

    def test_the_names_are_not_touched_in_arabic(self):
        self._doc({"data.name_one": "Layla"})
        self.assertIn("ليلى", self._page())

    # ---- عناصر القوايم
    def test_a_list_item_can_be_translated(self):
        """المفتاح بيتاخد من ‎translatable_entries‎ نفسها عشان الاختبار
        ما يعتمدش على اسم حقل ممكن يتغيّر."""
        rows = B.translatable_entries(self.inv.document)
        row = next(r for r in rows if r["value"] == "والد العروس")
        self.assertEqual(row["key"].count("."), 3)      # block.list.index.sub
        self._doc({row["key"]: "Father of the bride"})
        self.assertIn("Father of the bride", self._page("en"))

    # ---- الخطوط
    def test_english_can_take_its_own_fonts(self):
        """الخطوط العربية محارفها اللاتينية ناقصة أو وحشة."""
        doc = self.inv.document
        doc["theme"]["font_heading_en"] = "'Playfair Display', serif"
        doc["i18n"] = {"en": {self._hero_key(): "Wedding invitation"}}
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])
        self.assertIn("Playfair Display", self._page("en"))

    def test_an_empty_english_font_keeps_the_arabic_one(self):
        self._doc({self._hero_key(): "Wedding invitation"})
        self.assertIn("Amiri", self._page("en"))

    # ---- الأصل ما بيتلمسش
    def test_the_stored_document_is_never_rewritten(self):
        """‎apply_i18n‎ بيرجّع نسخة — لو عدّل الأصل كانت الترجمة هتاكل
        النص العربي من قاعدة البيانات."""
        doc = self._doc({self._hero_key(): "Wedding invitation"})
        B.apply_i18n(doc, "en")
        hero = next(b for b in doc["blocks"] if b["type"] == "hero")
        self.assertEqual(hero["props"]["kicker"], "دعوة زفاف")


# ==========================================================================
class TranslationKeyContractTests(BaseAppTest):
    """مفاتيح الجدول بيبنيها الجافاسكربت وبتقراها بايثون.

    أي فرق بينهم معناه ترجمة مكتوبة ومش ظاهرة، والمستخدم مش هيعرف ليه —
    فالاختبار ده بيقفل العقد بين الاتنين.
    """

    def setUp(self):
        super().setUp()
        self.js = (Path(settings.BASE_DIR) / "static/js/editor.js").read_text("utf-8")

    def test_python_builds_the_keys_it_can_read(self):
        rows = B.translatable_entries(self.inv.document, {"name_one": "ليلى"})
        table = {r["key"]: "EN " + r["key"] for r in rows}
        doc = dict(self.inv.document)
        doc["i18n"] = {"en": table}
        out = B.apply_i18n(doc, "en")
        # كل مفتاح بلوك اتطبّق فعلاً — مش اتجاهل بصمت
        hero = next(b for b in out["blocks"] if b["type"] == "hero")
        self.assertTrue(hero["props"]["kicker"].startswith("EN "))

    def test_the_js_uses_the_same_shapes(self):
        for shape in ('"data." + d[0]', '"settings." + s.key',
                      'block.id + "." + f.key',
                      'block.id + "." + f.key + "." + i + "." + sub.key'):
            with self.subTest(shape=shape):
                self.assertIn(shape, self.js)

    def test_the_js_only_offers_text_fields(self):
        self.assertIn("var I18N_TEXT = { text: 1, textarea: 1 }", self.js)
        self.assertEqual(B.TRANSLATABLE_TYPES, {"text", "textarea"})

    def test_dead_keys_are_swept_when_the_pane_opens(self):
        """قسم اتحذف كانت ترجمته بتفضل وتخلي زرار اللغة يظهر على الفاضي."""
        i = self.js.index("function renderI18nPane(")
        self.assertIn("delete table[k]", self.js[i:i + 2000])


# ==========================================================================
class CustomSectionTextTests(TestCase):
    """نصوص القسم المستورد — استخراجها من الكود واستبدالها فيه."""

    HTML = (
        '<section class="c">\n'
        '  <style>.x{color:red}</style>\n'
        '  <h1 data-move="el-1" class="t">YOU&#39;RE INVITED</h1>\n'
        '  <p data-move="el-2">Mohamed &amp; Farah</p>\n'
        '  <div data-move="el-3"><span>mixed</span> markup</div>\n'
        '  <img data-move="el-4" src="a.png">\n'
        '</section>'
    )

    def test_plain_text_units_are_found(self):
        keys = [m for m, _t in customtext.text_units(self.HTML)]
        self.assertEqual(keys, ["el-1", "el-2"])

    def test_the_text_comes_out_decoded(self):
        units = dict(customtext.text_units(self.HTML))
        self.assertEqual(units["el-2"], "Mohamed & Farah")

    def test_mixed_markup_is_left_alone(self):
        """الاستبدال كان هيمسح الوسم اللي جوّه والمصمّم مش هيفهم راح فين."""
        self.assertNotIn("el-3", dict(customtext.text_units(self.HTML)))

    def test_void_tags_are_not_offered(self):
        self.assertNotIn("el-4", dict(customtext.text_units(self.HTML)))

    def test_style_and_script_are_never_offered(self):
        """ده كان الباج: الجدول بيعرض CSS بدل الكلام."""
        for _m, text in customtext.text_units(self.HTML):
            self.assertNotIn("color:red", text)

    def test_broken_html_does_not_explode(self):
        self.assertIsInstance(customtext.text_units("<div><p data-move=x>hi"), list)

    # ---- الاستبدال
    def test_only_the_targeted_text_changes(self):
        out = customtext.replace_texts(self.HTML, {"el-1": "أنتم مدعوون"})
        self.assertIn("أنتم مدعوون", out)
        self.assertIn("Mohamed &amp; Farah", out)

    def test_the_tags_survive_byte_for_byte(self):
        """إعادة بناء الكود كانت هتغيّر الاقتباسات وتكسر ستايل القالب."""
        out = customtext.replace_texts(self.HTML, {"el-1": "X"})
        self.assertIn('<h1 data-move="el-1" class="t">', out)
        self.assertIn('<img data-move="el-4" src="a.png">', out)
        self.assertIn("<style>.x{color:red}</style>", out)

    def test_the_new_text_is_escaped(self):
        out = customtext.replace_texts(self.HTML, {"el-1": '<script>x</script>'})
        self.assertNotIn("<script>x", out)
        self.assertIn("&lt;script&gt;", out)

    def test_an_empty_map_returns_the_original(self):
        self.assertEqual(customtext.replace_texts(self.HTML, {}), self.HTML)


# ==========================================================================
class CustomSectionTranslationTests(BaseAppTest):
    """الترجمة جوّه قسم مستورد — من جدول الترجمة لحد صفحة الضيف."""

    HTML = '<div><h2 data-move="el-1">أهلاً بيكم</h2>' \
           '<p data-move="el-2">في فرحنا</p></div>'

    def setUp(self):
        super().setUp()
        doc = B.normalize_document({"blocks": [
            {"type": "custom_html", "id": "custom-1",
             "props": {"html": self.HTML, "css": ".a{color:red}"}},
        ]})
        self.inv.document = doc
        self.inv.save(update_fields=["document"])

    def _rows(self):
        return B.translatable_entries(self.inv.document)

    def _page(self, lang=""):
        url = self.inv.get_absolute_url() + (f"?lang={lang}" if lang else "")
        return self.client.get(url).content.decode()

    def test_the_words_show_up_not_the_code(self):
        values = [r["value"] for r in self._rows()]
        self.assertIn("أهلاً بيكم", values)
        self.assertIn("في فرحنا", values)

    def test_the_css_field_is_gone_from_the_table(self):
        """ده اللي كان بيملا الجدول بستايل بدل كلام."""
        self.assertNotIn(".a{color:red}", [r["value"] for r in self._rows()])
        self.assertFalse(any(r["key"].endswith(".css") for r in self._rows()))

    def test_the_key_points_inside_the_html(self):
        row = next(r for r in self._rows() if r["value"] == "أهلاً بيكم")
        self.assertEqual(row["key"], "custom-1.html#el-1")

    def test_translating_one_line_reaches_the_guest(self):
        doc = self.inv.document
        doc["i18n"] = {"en": {"custom-1.html#el-1": "Welcome"}}
        self.inv.document = B.normalize_document(doc)
        self.inv.save(update_fields=["document"])
        body = self._page("en")
        self.assertIn("Welcome", body)
        self.assertIn("في فرحنا", body)          # اللي مترجمش يفضل عربي

    def test_the_stored_html_is_never_rewritten(self):
        doc = self.inv.document
        doc["i18n"] = {"en": {"custom-1.html#el-1": "Welcome"}}
        B.apply_i18n(doc, "en")
        self.assertIn("أهلاً بيكم", doc["blocks"][0]["props"]["html"])

    def test_the_editor_reads_the_same_keys(self):
        js = (Path(settings.BASE_DIR) / "static/js/editor.js").read_text("utf-8")
        self.assertIn('block.id + "." + f.key + "#" + n.getAttribute("data-move")', js)
        self.assertIn("f.translate === false", js)


# ==========================================================================
class FaststartTests(TestCase):
    """نقل فهرس MP4 (moov) للأول — من غير ffmpeg.

    المتصفح مايقدرش يبدأ العرض قبل ما يقرا الـmoov. أغلب برامج التصدير
    بتكتبه في آخر الملف، فالمتصفح بينزّل الملف كله الأول. على اللوكال ده
    مالوش أثر ملحوظ، وعلى استضافة بطيئة بيبقى ثواني سودا قبل التشغيل —
    وده كان بيعدّي من غير أي رسالة لما ffmpeg مش متثبّت.
    """

    def _mp4(self, order=("ftyp", "mdat", "moov")):
        """ملف MP4 مبسّط بجدول ‎stco‎ حقيقي عشان نتأكد من تصحيح الأوفستات."""
        def box(kind, payload):
            return (len(payload) + 8).to_bytes(4, "big") + kind + payload

        ftyp = box(b"ftyp", b"isom" + b"\x00" * 8)
        media = b"\xAA" * 400
        # ترتيب الملف: ftyp ثم mdat — فبيانات mdat بتبدأ بعد ترويستها
        mdat_start = len(ftyp) + 8
        stco = box(b"stco", b"\x00\x00\x00\x00" + (1).to_bytes(4, "big")
                   + mdat_start.to_bytes(4, "big"))
        stbl = box(b"stbl", stco)
        minf = box(b"minf", stbl)
        mdia = box(b"mdia", minf)
        trak = box(b"trak", mdia)
        moov = box(b"moov", trak)
        mdat = box(b"mdat", media)
        parts = {"ftyp": ftyp, "mdat": mdat, "moov": moov}
        return b"".join(parts[k] for k in order), len(moov), mdat_start

    def test_it_spots_a_late_index(self):
        data, _m, _s = self._mp4(("ftyp", "mdat", "moov"))
        self.assertTrue(video.moov_is_late(data))

    def test_a_ready_file_is_left_alone(self):
        data, _m, _s = self._mp4(("ftyp", "moov", "mdat"))
        self.assertFalse(video.moov_is_late(data))
        self.assertIsNone(video.faststart_bytes(data))

    def test_the_index_moves_to_the_front(self):
        data, _m, _s = self._mp4()
        out = video.faststart_bytes(data)
        self.assertIsNotNone(out)
        self.assertFalse(video.moov_is_late(out))

    def test_nothing_is_re_encoded(self):
        """نفس البايتات بترتيب مختلف — الجودة والحجم زي ما هم."""
        data, _m, _s = self._mp4()
        out = video.faststart_bytes(data)
        self.assertEqual(len(out), len(data))
        self.assertIn(b"\xAA" * 400, out)

    def test_the_chunk_offsets_follow_the_move(self):
        """من غير التصحيح ده الفيديو بيبقى موجود ومش بيتفك."""
        data, moov_len, old_start = self._mp4()
        out = video.faststart_bytes(data)
        i = out.index(b"stco")
        moved = int.from_bytes(out[i + 4 + 8:i + 4 + 12], "big")
        self.assertEqual(moved, old_start + moov_len)
        # والقيمة الجديدة بتشاور فعلاً على أول بايت من بيانات mdat
        self.assertEqual(out[moved:moved + 4], b"\xAA" * 4)

    def test_junk_never_raises(self):
        for junk in (b"", b"not an mp4 at all", b"\x00\x00\x00\x08free"):
            with self.subTest(junk=junk[:12]):
                self.assertIsNone(video.faststart_bytes(junk))
                self.assertFalse(video.moov_is_late(junk))

    def test_a_truncated_file_is_left_alone(self):
        data, _m, _s = self._mp4()
        self.assertIsNone(video.faststart_bytes(data[:len(data) // 2]))

    # ---- المسار الكامل من غير ffmpeg
    def test_upload_is_fixed_even_without_ffmpeg(self):
        data, _m, _s = self._mp4()
        real = video.available
        video.available = lambda: False
        try:
            up = SimpleUploadedFile("clip.mp4", data, content_type="video/mp4")
            out, _secs = video.prepare_for_stream(up)
            self.assertFalse(video.moov_is_late(out.read()))
        finally:
            video.available = real

    def test_a_non_mp4_is_not_touched(self):
        real = video.available
        video.available = lambda: False
        try:
            up = SimpleUploadedFile("clip.webm", b"webm-ish", content_type="video/webm")
            out, _secs = video.prepare_for_stream(up)
            self.assertEqual(out.read(), b"webm-ish")
        finally:
            video.available = real


# ==========================================================================
class TextStyleGearTests(TestCase):
    """ترس النص — كل حقل نص في المحرر بياخد تنسيقه هو.

    قبل كده كان فيه ٤ بلوكات بس عندها حقول خط/حجم، بأسماء مالهاش علاقة
    بأسماء حقول النص (‎quote_font‎ لحقل اسمه ‎text‎)، ومفيش ولا حقل لون
    نص واحد. والترس كان لأربع حقول في الافتتاحية بخريطة مكتوبة بالإيد
    في ‎editor.js‎.
    """

    TEXT_TYPES = {"text", "textarea", "html"}

    def _owners(self, btype, spec):
        """حقول النص اللي المفروض ليها ترس في البلوك ده."""
        return [
            f for f in spec["props"]
            if f["type"] in self.TEXT_TYPES
            and not f.get("editor_hidden")
            and f.get("translate") is not False
            and (btype, f["key"]) not in B.NO_TEXT_STYLE
        ]

    # ---- المخطط
    def test_every_text_field_has_style_children(self):
        for btype, spec in B.BLOCK_REGISTRY.items():
            children = {}
            for f in spec["props"]:
                if f.get("style_of"):
                    children.setdefault(f["style_of"], set()).add(f["style_role"])
            for owner in self._owners(btype, spec):
                with self.subTest(block=btype, field=owner["key"]):
                    roles = children.get(owner["key"], set())
                    for role in ("font", "color", "size"):
                        self.assertIn(role, roles)

    def test_style_children_never_show_in_the_flat_list(self):
        """لو ظهروا في القايمة، يبقى نفس الإعداد في مكانين."""
        for btype, spec in B.BLOCK_REGISTRY.items():
            for f in spec["props"]:
                if f.get("style_of"):
                    with self.subTest(block=btype, field=f["key"]):
                        self.assertTrue(f.get("editor_hidden"))

    def test_the_old_font_fields_moved_into_the_gear_instead_of_doubling(self):
        """‎heading_font‎ و‎quote_size‎ و‎name_font‎ اتنقلوا مش اتكرروا."""
        cases = {
            "text": {"heading_font": "heading", "heading_size": "heading",
                     "body_size": "body", "body_line_height": "body"},
            "quote": {"quote_font": "text", "quote_size": "text"},
            "hero": {"name_font": "name_one", "name_size": "name_one",
                     "name_spacing": "name_one"},
        }
        for btype, expected in cases.items():
            fields = {f["key"]: f for f in B.BLOCK_REGISTRY[btype]["props"]}
            for key, owner in expected.items():
                with self.subTest(block=btype, field=key):
                    self.assertEqual(fields[key].get("style_of"), owner)
                    self.assertTrue(fields[key].get("editor_hidden"))
            # ومفيش مفتاح مولّد بيكرّر نفس الدور
            self.assertNotIn(f"{B.TEXT_STYLE_PREFIX}heading_font", fields)

    def test_intro_texts_keep_their_own_keys(self):
        """قيم الافتتاحية المحفوظة مالهاش تتغيّر — الترس بيقرا نفس المفاتيح."""
        fields = {f["key"]: f for f in B.SETTINGS_FIELDS}
        self.assertEqual(fields["intro_text_font"].get("style_of"), "intro_text")
        self.assertEqual(fields["intro_play_size"].get("style_of"),
                         "intro_play_label")

    def test_generated_font_fields_do_not_ship_their_own_font_list(self):
        """قايمة الخطوط بتتبعت مرة واحدة في المخطط، مش مع كل حقل."""
        heavy = [
            f["key"] for spec in B.BLOCK_REGISTRY.values() for f in spec["props"]
            if f["key"].startswith(B.TEXT_STYLE_PREFIX) and f.get("options")
            and f["type"] == "font"
        ]
        self.assertEqual(heavy, [])

    # ---- القوالب
    def test_every_gear_owner_has_a_hook_in_its_template(self):
        """من غير ‎data-ts‎ الترس بيتفتح ومايغيّرش حاجة على الشاشة."""
        root = Path(settings.BASE_DIR) / "templates" / "blocks"
        for btype, spec in B.BLOCK_REGISTRY.items():
            path = root / f"{btype}.html"
            if not path.exists():
                continue
            hooks = set(re.findall(r'data-ts="([^"]+)"', path.read_text("utf-8")))
            owners = {f["style_of"] for f in spec["props"] if f.get("style_of")}
            with self.subTest(block=btype):
                self.assertEqual(owners - hooks, set())

    # ---- الناتج
    def _hero(self, **props):
        block = B.make_block("hero")
        block["id"] = "hero-test1"
        block["props"].update(props)
        return block

    def test_a_set_value_becomes_one_scoped_rule(self):
        from system.renderer import text_style_css
        css = text_style_css([self._hero(
            ts_kicker_color="#b8914f", ts_kicker_weight="700",
            ts_kicker_align="center", ts_kicker_ls=1.5, ts_kicker_lh=1.4,
        )], {"max_width": 720})
        self.assertIn('[data-ts="kicker"]', css)
        for want in ("color:#b8914f", "font-weight:700", "text-align:center",
                     "letter-spacing:1.5px", "line-height:1.4"):
            self.assertIn(want, css)

    def test_the_size_follows_the_screen_like_the_rest(self):
        from system.renderer import text_style_css
        css = text_style_css([self._hero(ts_kicker_size=22)], {"max_width": 720})
        self.assertIn("font-size:clamp(", css)

    def test_an_untouched_document_adds_no_css_at_all(self):
        from system.renderer import text_style_css
        self.assertEqual(text_style_css([self._hero()], {"max_width": 720}), "")

    def test_junk_values_never_reach_the_stylesheet(self):
        from system.renderer import text_style_css
        css = text_style_css([self._hero(
            ts_kicker_color="red;background:url(javascript:alert(1))",
            ts_kicker_font="</style><script>x</script>",
            ts_kicker_weight="999", ts_kicker_align="justify",
            ts_kicker_size=9999,
        )], {"max_width": 720})
        self.assertEqual(css, "")

    def test_the_rule_still_finds_the_rsvp_section(self):
        """قالب تأكيد الحضور بيكتب ‎id="rsvp"‎ ثابت مش ‎block.id‎."""
        from system.renderer import text_style_css
        block = B.make_block("rsvp")
        block["id"] = "rsvp-test1"
        block["props"]["ts_submit_label_color"] = "#ffffff"
        css = text_style_css([block], {"max_width": 720})
        self.assertIn('[data-block="rsvp-test1"]', css)
        # ‎:is()‎ بتدّي القاعدة قوة المُعرِّف حتى وهي بتطابق بالسمة
        self.assertTrue(css.startswith(":is(#rsvp-test1,"))

    def test_the_stylesheet_reaches_the_page_with_the_offsets(self):
        doc = B.normalize_document({"blocks": [self._hero(ts_kicker_color="#123456")]})
        out = render_document(doc, editable=False)
        self.assertIn('[data-ts="kicker"]', out["layout_css"])

    # ---- المحرر
    def test_the_editor_builds_the_gear_from_the_schema(self):
        js = (Path(settings.BASE_DIR) / "static/js/editor.js").read_text("utf-8")
        self.assertIn("function buildTextStyleGear(", js)
        self.assertIn('specs[i].style_of === key', js)
        # الخريطة المكتوبة بالإيد اتشالت — حقل جديد في بايثون بيبان لوحده
        self.assertNotIn("INTRO_TEXT_FONT_FIELDS", js)

    def test_the_editor_rewrites_the_head_css_after_every_preview(self):
        """‎applyPreview‎ بتبدّل المسرح بس — من غير ده التعديل مايبانش."""
        js = (Path(settings.BASE_DIR) / "static/js/editor.js").read_text("utf-8")
        self.assertIn("function applyLayoutCss(", js)
        self.assertIn("applyLayoutCss(fdoc, data.layoutCss)", js)
        render = (Path(settings.BASE_DIR) / "templates/invitations/render.html"
                  ).read_text("utf-8")
        self.assertIn("<style data-lb-layout-css>", render)


# ==========================================================================
class OverlayTextGearTests(TestCase):
    """نص «فوق القسم» نص زي أي نص — لازم يبقى له ترس هو كمان.

    ده حقل ‎list‎، وعناصر القايمة ليها مسار تاني في ‎editor.js‎ ماكانش
    بيمرّر للترس حاجة، فحقول التنسيق كانت بتتفرد تحت النص.
    """

    def _sub(self, key):
        return next(f for f in B.SECTION_TEXT_OVERLAY_FIELD["fields"]
                    if f["key"] == key)

    def test_the_style_fields_belong_to_the_text(self):
        for key, role in (("font", "font"), ("color", "color"), ("size", "size")):
            with self.subTest(key=key):
                spec = self._sub(key)
                self.assertEqual(spec["style_of"], "text")
                self.assertEqual(spec["style_role"], role)
                self.assertTrue(spec["editor_hidden"])

    def test_position_fields_stay_out_in_the_open(self):
        """العرض والموضع مش تنسيق — ليهم مقابض في المعاينة."""
        for key in ("width", "x", "y"):
            with self.subTest(key=key):
                self.assertIsNone(self._sub(key).get("style_of"))
                self.assertFalse(self._sub(key).get("editor_hidden"))

    def test_the_saved_keys_did_not_move(self):
        """نصوص محفوظة قبل كده لازم تفضل شغالة بنفس المفاتيح."""
        keys = [f["key"] for f in B.SECTION_TEXT_OVERLAY_FIELD["fields"]]
        self.assertEqual(keys[:3], ["text", "color", "font"])

    def test_a_size_reaches_the_page_as_a_fluid_value(self):
        style = invite_tags.video_text_style(
            {"text": "x", "size": 40, "color": "#ffffff"})
        self.assertIn("--section-text-size:clamp(", style)

    def test_zero_size_writes_nothing(self):
        """صفر = حجم القالب زي ما هو، مش ‎0px‎."""
        style = invite_tags.video_text_style({"text": "x", "size": 0})
        self.assertNotIn("--section-text-size", style)

    def test_junk_size_is_dropped(self):
        for bad in ("9999", "-5", "abc", None):
            with self.subTest(size=bad):
                self.assertNotIn(
                    "--section-text-size:clamp(1",
                    invite_tags.video_text_style({"text": "x", "size": bad}))

    def test_the_stylesheet_reads_the_variable(self):
        css = (Path(settings.BASE_DIR) / "static/css/invite.css").read_text("utf-8")
        self.assertIn("font-size: var(--section-text-size, inherit)", css)

    def test_the_editor_passes_the_gear_into_list_items(self):
        js = (Path(settings.BASE_DIR) / "static/js/editor.js").read_text("utf-8")
        self.assertIn("var itemCtx = {", js)
        self.assertIn("itemSet,\n                itemCtx", js.replace("\r\n", "\n"))
        # ومابيفردش الحقول المخفية تحت النص
        self.assertIn("if (sub.editor_hidden) return;", js)
