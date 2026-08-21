"""تصدير كشف الضيوف كملف إكسل ومعاه رموز QR.

الملف ده بيتبعت للقاعة عشان تعرف مين المدعوّين وكام واحد مع كل ضيف،
وتقدر تمسح الرمز من الشاشة أو تطبعه.

كل صف فيه صورة QR فعلية مش رابط — القاعة غالباً مالهاش نت على الباب.
"""

from __future__ import annotations

import io

from . import qrcodes

# QR أصغر من بتاع التحميل — الصف في إكسل مش هيستحمل حجم كبير،
# و٦ بكسل للمربّع كفاية للمسح من الشاشة أو ورق A4.
QR_BOX = 5
ROW_HEIGHT = 92          # نقطة — بيسع صورة ١٢٠px تقريباً
QR_COL_WIDTH = 18

HEADERS = [
    ("الكود", 16),
    ("النوع", 12),
    ("الضيف", 26),
    ("الهاتف", 18),
    ("الحالة", 12),
    ("مسموح", 10),
    ("دخل", 8),
    ("متبقي", 9),
    ("الرابط الفردي", 46),
    ("QR", QR_COL_WIDTH),
]


def build(invitation, guests, *, base_url: str) -> bytes:
    """يرجّع محتوى ملف .xlsx كـbytes."""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "الضيوف"
    ws.sheet_view.rightToLeft = True          # الكشف عربي

    title_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    head_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=11)
    mono_font = Font(name="Arial", size=10)
    head_fill = PatternFill("solid", fgColor="8A6A33")
    title_fill = PatternFill("solid", fgColor="2C2620")
    thin = Side(style="thin", color="D8CDBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # ---- ترويسة الملف
    last_col = get_column_letter(len(HEADERS))
    ws.merge_cells(f"A1:{last_col}1")
    cell = ws["A1"]
    names = " و ".join(n for n in [invitation.name_one, invitation.name_two] if n)
    cell.value = f"كشف ضيوف {names or invitation.title}"
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = center
    ws.row_dimensions[1].height = 30

    for i, (label, width) in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=i, value=label)
        c.font = head_font
        c.fill = head_fill
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    status_labels = {"active": "نشط", "used": "مستخدم", "none": "بدون تصريح"}

    row = 3
    for guest in guests:
        link = base_url.rstrip("/") + guest.get_absolute_url()
        values = [
            guest.pass_code,
            guest.get_source_display(),
            guest.name,
            guest.phone or "—",
            status_labels.get(guest.pass_status, ""),
            guest.entries_allowed,
            guest.entries_used,
            guest.entries_left,
            link,
            "",
        ]
        for i, value in enumerate(values, start=1):
            c = ws.cell(row=row, column=i, value=value)
            c.font = mono_font if i in (1, 9) else body_font
            c.alignment = right if i in (3, 9) else center
            c.border = border

        png = qrcodes.png_for(link, label=guest.pass_code, box_size=QR_BOX)
        if png:
            img = XLImage(io.BytesIO(png))
            # نثبّت الارتفاع ونسيب العرض يتظبط بنفس النسبة
            ratio = img.width / img.height if img.height else 1
            img.height = 110
            img.width = int(110 * ratio)
            img.anchor = f"{get_column_letter(len(HEADERS))}{row}"
            ws.add_image(img)
        ws.row_dimensions[row].height = ROW_HEIGHT
        row += 1

    if row == 3:
        ws.cell(row=3, column=1, value="مفيش ضيوف في الكشف ده.").font = body_font

    # ---- ورقة الملخّص
    s2 = wb.create_sheet("ملخّص")
    s2.sheet_view.rightToLeft = True
    s2.column_dimensions["A"].width = 30
    s2.column_dimensions["B"].width = 14
    rows = [
        ("عدد الضيوف", len(guests)),
        ("إجمالي الدخلات المسموحة", sum(g.entries_allowed for g in guests)),
        ("اللي دخلوا فعلاً", sum(g.entries_used for g in guests)),
        ("المتبقي", sum(g.entries_left for g in guests)),
        ("تصاريح نشطة", sum(1 for g in guests if g.pass_status == "active")),
        ("تصاريح مستخدمة", sum(1 for g in guests if g.pass_status == "used")),
    ]
    for i, (label, value) in enumerate(rows, start=1):
        a = s2.cell(row=i, column=1, value=label)
        a.font = Font(name="Arial", size=11, bold=True)
        a.alignment = right
        b = s2.cell(row=i, column=2, value=value)
        b.font = body_font
        b.alignment = center

    note = s2.cell(row=len(rows) + 2, column=1,
                   value="الأرقام دي لحظة التصدير — مش بتتحدّث لوحدها.")
    note.font = Font(name="Arial", size=10, italic=True, color="7B6F62")
    note.alignment = right

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
